#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Producción - Paso 2: Generador de Sitios Web Borrador Premium (CDMX)
Aplica: SEO Avanzado (JSON-LD, OG), Animaciones CSS/JS, Paletas de Marca Personalizadas,
Formato WebP/Imágenes Optimizadas, Botones de WhatsApp CRO y FAQs Interactivas.
"""

import os
import re
import sys
import shutil
import subprocess
import openpyxl
from datetime import datetime

BASE_DIR = '/Users/beto/Library/CloudStorage/OneDrive-Personal(2)/Freelancer/clientes/web'
EXCEL_PATH = os.path.join(BASE_DIR, 'Clientes.xlsx')

def slugify(text):
    text = text.lower().strip()
    # Reemplazar acentos
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n', 'ü': 'u'}
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text.strip('-')

# Definición de marcas y personalizaciones por empresa
BRANDS_CONFIG = {
    'Kintsu Dental': {
        'slug': 'kintsu-dental',
        'nicho': 'Dentist',
        'primary': '#0f766e', # Teal Japonés / Minimalista
        'accent': '#14b8a6',
        'bg_gradient': 'linear-gradient(135deg, #042f2e 0%, #0f766e 100%)',
        'hero_title': 'Odontología Estética y Bienestar Dental en CDMX',
        'hero_sub': 'Inspirados en la perfección del arte Kintsugi: restauramos y resaltamos la belleza natural de tu sonrisa con tecnología de vanguardia.',
        'hero_img': 'https://images.unsplash.com/photo-1629909613654-28e377c37b09?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'kintsu dental, clinica dental cdmx, odontologia estetica cdmx, diseno de sonrisa, dentista cdmx',
        'servicios': [
            ('Diseño de Sonrisa Kintsu', 'Carillas cerámicas de alta definición y estética dental personalizada para reflejar tu personalidad.'),
            ('Ortodoncia Invisible', 'Alineación dental discreta con alineadores transparentes de última generación.'),
            ('Blanqueamiento Láser', 'Tratamiento de aclarado dental rápido, seguro y sin sensibilidad molesta.'),
            ('Odontología Preventiva & Limpieza', 'Higiene ultrasonica profunda y remoción de sarro para proteger tu salud bucal.')
        ]
    },
    'Clínica Dental Amsterdent': {
        'slug': 'clinica-dental-amsterdent',
        'nicho': 'Dentist',
        'primary': '#0284c7', # Azul Condesa / Moderno
        'accent': '#38bdf8',
        'bg_gradient': 'linear-gradient(135deg, #0c4a6e 0%, #0284c7 100%)',
        'hero_title': 'Tu Clínica Dental de Confianza en La Condesa, CDMX',
        'hero_sub': 'Atención odontológica integral en un ambiente cómodo, moderno y relajado en el corazón de la Condesa.',
        'hero_img': 'https://images.unsplash.com/photo-1588776814546-1ffcf47267a5?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'amsterdent, clinica dental condesa, dentista condesa cdmx, ortodoncia condesa, implantes dentales cdmx',
        'servicios': [
            ('Implantes Dentales de Titanio', 'Restauración permanente de piezas dentales perdidas con máxima firmeza y estética natural.'),
            ('Ortodoncia & Brackets', 'Corrección de mordida y alineación con brackets estéticos y convencionales.'),
            ('Endodoncia Computarizada', 'Tratamientos de conducto indoloros con tecnología digital para salvar tus piezas naturales.'),
            ('Limpieza Dental Profunda', 'Eliminación de placa bacteriana y sarro con pulido de alta frecuencia.')
        ]
    },
    'KlinikDent México': {
        'slug': 'klinikdent-mexico',
        'nicho': 'Dentist',
        'primary': '#1e3a8a', # Azul Zafiro & Premium
        'accent': '#3b82f6',
        'bg_gradient': 'linear-gradient(135deg, #172554 0%, #1e3a8a 100%)',
        'hero_title': 'Odontología de Alta Especialidad en Del Valle, CDMX',
        'hero_sub': 'Combinamos experiencia médica de primer nivel con equipos digitales de diagnóstico preciso para cuidar tu sonrisa.',
        'hero_img': 'https://images.unsplash.com/photo-1606811841689-23dfddce3e95?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'klinikdent mexico, clinica dental del valle, odontologia especializada cdmx, blanqueamiento dental cdmx',
        'servicios': [
            ('Estética Dental & Carillas', 'Perfeccionamos la forma, color y simetría de tus dientes con materiales cerámicos duraderos.'),
            ('Cirugía Oral & Cordales', 'Extracción segura de muelas del juicio con técnicas mínimamente invasivas.'),
            ('Rehabilitación Oral Completa', 'Recuperación de la función masticatoria mediante prótesis fijas y removibles.'),
            ('Odontopediatría Especializada', 'Cuidado bucal amigable y sin estrés para los pequeños de la familia.')
        ]
    },
    'Dental Studio MX': {
        'slug': 'dental-studio-mx',
        'nicho': 'Dentist',
        'primary': '#7c3aed', # Violeta Innovación & Studio
        'accent': '#a855f7',
        'bg_gradient': 'linear-gradient(135deg, #3b0764 0%, #7c3aed 100%)',
        'hero_title': 'Diseño de Sonrisas & Ortodoncia Digital en CDMX',
        'hero_sub': 'Sedes en Narvarte, Polanco, Coyoacán y Pedregal. Tecnología 3D para transformar tu sonrisa sin complicaciones.',
        'hero_img': 'https://images.unsplash.com/photo-1598256989800-fe5f95da9787?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'dental studio mx, ortodoncia invisible narvarte, carillas polanco, clinica dental coyoacan, dentista pedregal',
        'servicios': [
            ('Escaneo Intraoral 3D', 'Diagnóstico digital rápido sin la molestia de las pastas tradicionales.'),
            ('Alineadores Invisibles Studio', 'Dientes rectos y alineados en meses sin que nadie note tus brackets.'),
            ('Diseño de Sonrisa Digital (DSD)', 'Visualiza el resultado final de tu sonrisa en pantalla antes de iniciar tu tratamiento.'),
            ('Profilaxis y Blanqueamiento', 'Higiene ultrasónica completa acompañada de sesión de blanqueamiento LED.')
        ]
    },
    'Clínica Dental Proboca': {
        'slug': 'clinica-dental-proboca',
        'nicho': 'Dentist',
        'primary': '#0d9488', # Turquesa Familiar
        'accent': '#2dd4bf',
        'bg_gradient': 'linear-gradient(135deg, #115e59 0%, #0d9488 100%)',
        'hero_title': 'Tu Salud Bucal en Buenas Manos — Av. Zaragoza, CDMX',
        'hero_sub': 'Atención odontológica accesible, profesional y honesta para toda la familia con los más altos estándares de higiene.',
        'hero_img': 'https://images.unsplash.com/photo-1579684385127-1ef15d508118?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'clinica dental proboca, dentista zaragoza cdmx, consulta dental accesible cdmx, obturaciones resina',
        'servicios': [
            ('Consultas y Diagnóstico con Cámara', 'Evaluación completa de tu boca con imágenes ampliadas en monitor.'),
            ('Resinas Estéticas del Color Dental', 'Eliminación de caries y restauración con empastes invisibles del mismo tono de tu diente.'),
            ('Extracciones y Alivio del Dolor', 'Atención oportuna de molestias e infecciones bucales.'),
            ('Limpiezas Dentales Preventivas', 'Remoción de sarro acumulado para prevenir la inflamación de encías (gingivitis).')
        ]
    },
    'Clínica Dental Rehabilitarte': {
        'slug': 'clinica-dental-rehabilitarte',
        'nicho': 'Dentist',
        'primary': '#0369a1', # Azul Océano Clínico
        'accent': '#0284c7',
        'bg_gradient': 'linear-gradient(135deg, #075985 0%, #0369a1 100%)',
        'hero_title': 'Rehabilitación Oral y Odontología Integral en Tlalpan',
        'hero_sub': 'Especialistas en devolver la función y estética a tu dentadura. Atención personalizada en La Joya, Tlalpan, CDMX.',
        'hero_img': 'https://images.unsplash.com/photo-1629909615184-74f495363b67?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'clinica dental rehabilitarte, dentista tlalpan cdmx, rehabilitacion oral la joya, protesis dentales tlalpan',
        'servicios': [
            ('Prótesis e Implantes Dentales', 'Soluciones fijas y confortables para reemplazar dientes faltantes y recuperar tu masticación.'),
            ('Tratamientos de Periodoncia', 'Cuidado especializado de encías sangrantes y soporte de los dientes.'),
            ('Endodoncia y Salvaguarda de Piezas', 'Limpieza profunda de conductos infectados para evitar extracciones innecesarias.'),
            ('Blanqueamiento & Estética', 'Aclarado dental profesional para lucir una dentadura juvenil y radiante.')
        ]
    },
    'Clínica Dental DEntAle': {
        'slug': 'clinica-dental-dentale',
        'nicho': 'Dentist',
        'primary': '#0284c7', # Azul Sky Moderno
        'accent': '#38bdf8',
        'bg_gradient': 'linear-gradient(135deg, #0369a1 0%, #0284c7 100%)',
        'hero_title': 'Excelencia Dental al Alcance de Tu Familia en Iztapalapa',
        'hero_sub': 'Ubicados en Col. Citlalli. Brindamos tratamientos modernos, confortables y duraderos con odontólogos capacitados.',
        'hero_img': 'https://images.unsplash.com/photo-1606811841689-23dfddce3e95?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'clinica dental dentale, dentista iztapalapa cdmx, ortodoncia citlalli, implantes dentales iztapalapa',
        'servicios': [
            ('Ortodoncia para Niños y Adultos', 'Alineación de dientes desalineados con brackets metálicos y estéticos.'),
            ('Restauraciones en Resina', 'Eliminación de picaduras de caries con resinas fotocurables de alta dureza.'),
            ('Tratamiento de Canales (Endodoncia)', 'Procedimientos precisos bajo anestesia local para quitar dolores de muela.'),
            ('Higiene Bucal Ultrasonica', 'Limpieza rápida con ultrasonido que remueve manchas y sarro difícil.')
        ]
    },
    'Grupo Sonríe Mx': {
        'slug': 'grupo-sonrie-mx',
        'nicho': 'Dentist',
        'primary': '#db2777', # Rosa/Magenta Sonrisa Radiante
        'accent': '#f43f5e',
        'bg_gradient': 'linear-gradient(135deg, #831843 0%, #db2777 100%)',
        'hero_title': 'Transformamos Tu Sonrisa con Pasión y Tecnología en CDMX',
        'hero_sub': 'Especialistas en odontología estética, carillas y diseño de sonrisa con resultados espectaculares y naturales.',
        'hero_img': 'https://images.unsplash.com/photo-1598256989800-fe5f95da9787?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'grupo sonrie mx, diseno de sonrisa cdmx, carillas de porcelana cdmx, blanqueamiento dental cdmx',
        'servicios': [
            ('Carillas de Porcelana & Lentes Cerámicos', 'Modificación de forma, tamaño y blanco radiante sin desgaste agresivo del diente.'),
            ('Blanqueamiento Dental VIP', 'Sesiones intensivas que aclaran hasta 4 tonos en una sola visita.'),
            ('Ortodoncia Invisible', 'Alineación estética con férulas transparentes hechas a tu medida.'),
            ('Diseño Digital de Sonrisa', 'Planeación computarizada para lograr la proporción áurea facial ideal.')
        ]
    },
    'DS Consultorio Dental': {
        'slug': 'ds-consultorio-dental',
        'nicho': 'Dentist',
        'primary': '#2563eb', # Azul Royal Clínico
        'accent': '#60a5fa',
        'bg_gradient': 'linear-gradient(135deg, #1e3a8a 0%, #2563eb 100%)',
        'hero_title': 'Atención Odontológica Integral & Personalizada en CDMX',
        'hero_sub': 'Comprometidos con tu salud oral. Ofrecemos valoraciones detalladas y tratamientos éticos al mejor costo.',
        'hero_img': 'https://images.unsplash.com/photo-1588776814546-1ffcf47267a5?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'ds consultorio dental, dentista cdmx, consulta dental cdmx, endodoncia resina curacion muela',
        'servicios': [
            ('Limpieza Dental Ultrasonica', 'Limpieza profunda sin dolor que deja tus dientes suaves y libres de sarro.'),
            ('Tratamiento de Caries con Resinas', 'Restauraciones duraderas e imperceptibles al sonreír.'),
            ('Guardas Oclusales contra Bruxismo', 'Protectores hechos a medida para evitar el desgaste nocturno por apretar los dientes.'),
            ('Extracciones Dentales Simples', 'Procedimientos seguros e higiénicos para piezas no restaurables.')
        ]
    },
    'Clínica Dental Godiental': {
        'slug': 'clinica-dental-godiental',
        'nicho': 'Dentist',
        'primary': '#0f766e', # Esmeralda Odontológico
        'accent': '#14b8a6',
        'bg_gradient': 'linear-gradient(135deg, #115e59 0%, #0f766e 100%)',
        'hero_title': 'Especialistas en Implantes & Cirugía Bucal en CDMX',
        'hero_sub': 'Recupera la firmeza y confianza de sonreír con tratamientos de implantología de última generación.',
        'hero_img': 'https://images.unsplash.com/photo-1629909613654-28e377c37b09?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'clinica dental godiental, implantes dentales cdmx, cirugia bucal cdmx, dentista implantes cdmx',
        'servicios': [
            ('Implantes Dentales Guiados', 'Reemplazo de raíz dental con pernos de titanio guiados por tomografía digital.'),
            ('Prótesis sobre Implantes', 'Fijación de puentes y dentaduras completas sobre implantes sin movilidad.'),
            ('Extracción de Terceros Molares', 'Retiro de cordales retenidas con mínima inflamación postoperatoria.'),
            ('Estética Bucal & Blanqueamiento', 'Complementa tu tratamiento de implantes con una sonrisa más blanca y rejuvenecida.')
        ]
    },
    'Rehavilita Fisioterapia': {
        'slug': 'rehavilita-fisioterapia',
        'nicho': 'PhysicalTherapy',
        'primary': '#14436c', # Navy Fisioterapia / Médico
        'accent': '#2ecfb4',
        'bg_gradient': 'linear-gradient(135deg, #0b2545 0%, #14436c 100%)',
        'hero_title': 'Centro de Fisioterapia & Terapia Ocupacional en Col. Del Valle',
        'hero_sub': 'Especialistas en terapia física, rehabilitación del lenguaje, respiratoria y masajes terapéuticos en Miguel Laurent 510, CDMX.',
        'hero_img': 'https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'rehavilita fisioterapia, fisioterapia del valle cdmx, terapia fisica miguel laurent, rehabilitacion respiratoria cdmx',
        'servicios': [
            ('Fisioterapia & Rehabilitación Física', 'Tratamiento integral de lesiones traumatológicas, deportivas y dolores de columna.'),
            ('Terapia de Lenguaje & Ocupacional', 'Rehabilitación adaptativa y cognitiva para niños, adultos y adultos mayores.'),
            ('Fisioterapia Respiratoria', 'Técnicas avanzadas para mejorar la capacidad pulmonar y despejar vías aéreas.'),
            ('Masajes Terapéuticos & Descontracturantes', 'Liberación de tensión muscular profunda, contracturas y estrés laboral.')
        ]
    },
    'Fisioterapia y Rehabilitación Vértiz': {
        'slug': 'fisioterapia-y-rehabilitacion-vertiz',
        'nicho': 'PhysicalTherapy',
        'primary': '#0284c7', # Azul Medicina Deportiva
        'accent': '#38bdf8',
        'bg_gradient': 'linear-gradient(135deg, #075985 0%, #0284c7 100%)',
        'hero_title': 'Medicina Deportiva & Fisioterapia Integral — Dr. Vertiz, CDMX',
        'hero_sub': 'Recupera tu movilidad sin dolor. Ubicados en Dr. José María Vértiz 1218, Benito Juárez, CDMX.',
        'hero_img': 'https://images.unsplash.com/photo-1519823551278-64ac92734fb1?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'fisioterapia vertiz cdmx, medicina deportiva benito juarez, rehabilitacion fisica vertiz, masaje terapeutico cdmx',
        'servicios': [
            ('Rehabilitación Deportiva de Alto Rendimiento', 'Tratamiento de esguinces, desgarros musculares y tendinitis en deportistas.'),
            ('Fisioterapia Traumatológica & Postquirúrgica', 'Recuperación asistida tras cirugías ortopédicas de rodilla, hombro y tobillo.'),
            ('Manejo del Dolor de Espalda & Ciática', 'Terapia manual y electroterapia para aliviar hernias discales y lumbalgias.'),
            ('Masaje Terapéutico Deportivo', 'Descarga muscular profunda para favorecer la circulación y prevención de lesiones.')
        ]
    },
    'SCA Clínica de Fisioterapia': {
        'slug': 'sca-clinica-de-fisioterapia',
        'nicho': 'PhysicalTherapy',
        'primary': '#0d9488', # Teal Deportivo / Clínico
        'accent': '#2dd4bf',
        'bg_gradient': 'linear-gradient(135deg, #115e59 0%, #0d9488 100%)',
        'hero_title': 'Clínica de Fisioterapia & Rehabilitación Física en Iztapalapa',
        'hero_sub': 'Samuel Gompers 34, Col. Jacarandas. Vuelve a moverte libremente con protocolos de fisioterapia avanzada.',
        'hero_img': 'https://images.unsplash.com/photo-1576091160550-2173dba999ef?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'sca clinica de fisioterapia, fisioterapia iztapalapa, rehabilitacion jacarandas cdmx, terapia fisica iztapalapa',
        'servicios': [
            ('Fisioterapia Ortopédica & Articular', 'Tratamiento de contracturas, lumbalgias, cervicalgias y desgaste articular (artrosis).'),
            ('Rehabilitación Física Deportiva', 'Recuperación de capacidad funcional para atletas y aficionados a la actividad física.'),
            ('Electroterapia & Ultrasonido Terapéutico', 'Tecnología analgésica y antiinflamatoria para acelerar la cicatrización muscular.'),
            ('Reeducación Postural & Ejercicio Guiado', 'Corrección de postura y fortalecimiento muscular adaptado a tu estilo de vida.')
        ]
    },
    'Clínica Fisioterapia Santillán': {
        'slug': 'clinica-fisioterapia-santillan',
        'nicho': 'PhysicalTherapy',
        'primary': '#1e3a8a', # Azul Zafiro Rehabilitación
        'accent': '#60a5fa',
        'bg_gradient': 'linear-gradient(135deg, #172554 0%, #1e3a8a 100%)',
        'hero_title': 'Especialistas en Tratamiento del Dolor & Rehabilitación en CDMX',
        'hero_sub': 'Atención profesional en Bajío 335, Roma Sur / Cuauhtémoc. Evaluación personalizada para erradicar el dolor de raíz.',
        'hero_img': 'https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'fisioterapia santillan cdmx, fisioterapia roma sur, clinica fisioterapia bajio cdmx, alivio dolor espalda cdmx',
        'servicios': [
            ('Tratamiento Integral del Dolor Crónico', 'Manejo especializado de lumbalgias, hernias de disco, fibromialgia y rigidez.'),
            ('Fisioterapia Postquirúrgica', 'Rehabilitación gradual tras reemplazos articulares, artroscopias y fijación de fracturas.'),
            ('Terapia Manual & Puntos Gatillo', 'Desactivación manual de nudos musculares dolientes e hipertonía muscular.'),
            ('Fisioterapia Preventiva & Ergonomía', 'Asesoría para evitar dolores posturales ocasionados por trabajo de oficina.')
        ]
    },
    'Podología & Fisioterapia Pies Alegres': {
        'slug': 'podologia-fisioterapia-pies-alegres',
        'nicho': 'PhysicalTherapy',
        'primary': '#0284c7', # Azul Turquesa Biomecánico
        'accent': '#38bdf8',
        'bg_gradient': 'linear-gradient(135deg, #075985 0%, #0284c7 100%)',
        'hero_title': 'Fisioterapia Podológica & Estudio Biomecánico del Pie en CDMX',
        'hero_sub': 'Eje 5 Sur Eugenia 926, Del Valle. Especialistas en análisis de marcha, fascitis plantar y salud integral del pie.',
        'hero_img': 'https://images.unsplash.com/photo-1519823551278-64ac92734fb1?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'pies alegres eugenia cdmx, fisioterapia podologica cdmx, estudio biomecanico pie eugenia, fascitis plantar cdmx',
        'servicios': [
            ('Fisioterapia para Fascitis Plantar & Espolón', 'Tratamiento analgésico y de estiramiento para dolor intenso en el talón al caminar.'),
            ('Estudio Biomecánico de la Pisada & Marcha', 'Análisis computarizado de presión plantar para detectar desviaciones de apoyo.'),
            ('Plantillas Ortopédicas Personalizadas', 'Diseño e impresión de plantillas a medida para corregir postura y aliviar rodillas/espalda.'),
            ('Rehabilitación de Tobillo & Pie', 'Recuperación de movilidad y estabilidad tras esguinces recurrentes de tobillo.')
        ]
    },
    'CERCARDIO Especialidades Médicas': {
        'slug': 'cercardio-especialidades-medicas',
        'nicho': 'MedicalClinic',
        'primary': '#b91c1c', # Rojo Médico Cardiovascular & Navy
        'accent': '#ef4444',
        'bg_gradient': 'linear-gradient(135deg, #7f1d1d 0%, #b91c1c 100%)',
        'hero_title': 'Centro de Especialidades & Rehabilitación Cardiovascular en CDMX',
        'hero_sub': 'Sedes en Hospital Ángeles Lindavista y Hospital Boutique Riobamba. Diagnóstico y cuidado del corazón con médicos certificados.',
        'hero_img': 'https://images.unsplash.com/photo-1516549655169-df83a0774514?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'cercardio cdmx, cardiologo angeles lindavista, rehabilitacion cardiovascular cdmx, checkup cardiaco cdmx',
        'servicios': [
            ('Checkup Cardiológico Completo', 'Electrocardiograma, ecocardiograma doppler y prueba de esfuerzo para evaluar tu salud cardíaca.'),
            ('Programa de Rehabilitación Cardiovascular', 'Entrenamiento físico médico supervisado tras infarto o cirugías de corazón.'),
            ('Monitoreo Holter & Presión Arterial', 'Seguimiento continuo de 24 hrs para arritmias e hipertensión arterial.'),
            ('Consulta de Especialidades Médicas', 'Atención integral por cardiólogos, internistas y especialistas quirúrgicos.')
        ]
    },
    'Seishi Centro de Atención Psicológica': {
        'slug': 'seishi-centro-atencion-psicologica',
        'nicho': 'MedicalClinic',
        'primary': '#0f766e', # Verde Menta & Calma
        'accent': '#2dd4bf',
        'bg_gradient': 'linear-gradient(135deg, #134e4a 0%, #0f766e 100%)',
        'hero_title': 'Centro de Atención Psicológica & Bienestar Emocional en CDMX',
        'hero_sub': 'Un espacio seguro, ético y profesional para sanar, crecer y recuperar el equilibrio mental en tu vida.',
        'hero_img': 'https://images.unsplash.com/photo-1573496359142-b8d87734a5a2?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'seishi centro de atencion psicologica, psicologo cdmx, terapia de pareja cdmx, psicoterapia ansiedad cdmx',
        'servicios': [
            ('Psicoterapia Individual para Adultos', 'Acompañamiento profesional para superar ansiedad, depresión, duelo y estrés.'),
            ('Terapia de Pareja & Relaciones', 'Herramientas de comunicación asertiva para resolver conflictos y fortalecer vínculos.'),
            ('Terapia Infantil & Adolescentes', 'Orientación emocional y apoyo psicopedagógico para niños y jóvenes.'),
            ('Talleres de Desarrollo Personal & Mindfulness', 'Espacios grupales para el manejo de emociones y autocuidado.')
        ]
    },
    'Centro Bewusst Machen': {
        'slug': 'centro-bewusst-machen',
        'nicho': 'MedicalClinic',
        'primary': '#6d28d9', # Púrpura Salud Mental & Reflexión
        'accent': '#8b5cf6',
        'bg_gradient': 'linear-gradient(135deg, #4c1d95 0%, #6d28d9 100%)',
        'hero_title': 'Salud Mental Integral & Psicoterapia Especializada en CDMX',
        'hero_sub': 'Bewusst Machen: "Hacer consciente". Te ayudamos a comprender tus procesos emocionales y transformar tu calidad de vida.',
        'hero_img': 'https://images.unsplash.com/photo-1527689368864-3a821dbccc34?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'centro bewusst machen, salud mental cdmx, psicoterapia integral cdmx, psicologo especialista cdmx',
        'servicios': [
            ('Psicoterapia Cognitivo-Conductual', 'Tratamiento focalizado para reestructurar pensamientos obsesivos, fobias y ansiedad.'),
            ('Orientación Psicológica & Coaching', 'Estrategias breves para toma de decisiones, liderazgo y metas personales.'),
            ('Evaluación & Psicodiagnóstico Clínico', 'Pruebas psicométricas completas para un diagnóstico claro y certero.'),
            ('Terapia Online & Presencial en CDMX', 'Modalidades flexibles para adaptarnos a tus horarios y disponibilidad.')
        ]
    },
    'Clínica Médica Palmas - Dra. Yanina Rubio': {
        'slug': 'clinica-medica-palmas-dra-yanina-rubio',
        'nicho': 'MedicalClinic',
        'primary': '#0369a1', # Azul Salud Integral / Palmas
        'accent': '#38bdf8',
        'bg_gradient': 'linear-gradient(135deg, #0c4a6e 0%, #0369a1 100%)',
        'hero_title': 'Atención Psicológica & Salud Mental en Clínica Médica Palmas, CDMX',
        'hero_sub': 'Consulta privada por la Lic. Yanina Rubio Manzo. Atención personalizada en un entorno médico distinguido y confidencial.',
        'hero_img': 'https://images.unsplash.com/photo-1551076805-e1869033e561?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'clinica medica palmas cdmx, psicologa yanina rubio, consulta psicologica palmas, psicoterapia lomas cdmx',
        'servicios': [
            ('Psicoterapia Individual de Alta Especialidad', 'Abordaje clínico confidencial para el manejo del estrés ejecutivo, depresión y traumas.'),
            ('Consulta en Clínica Médica Especializada', 'Espacio médico acondicionado con máxima privacidad y confort.'),
            ('Manejo de Burnout & Estrés Laboral', 'Terapia orientada a profesionales y ejecutivos con alta exigencia laboral.'),
            ('Acompañamiento en Procesos de Cambio', 'Apoyo psicológico para afrontar divorcios, pérdidas o transiciones de vida importantes.')
        ]
    },
    'Vitalmente Centro Médico & Psicología': {
        'slug': 'vitalmente-centro-medico-psicologia',
        'nicho': 'MedicalClinic',
        'primary': '#0d9488', # Verde Vital / Esmeralda
        'accent': '#14b8a6',
        'bg_gradient': 'linear-gradient(135deg, #115e59 0%, #0d9488 100%)',
        'hero_title': 'Centro de Salud Integral & Psicología Clínica en Lomas, CDMX',
        'hero_sub': 'Calle Sierra de Roraima 15, CDMX. Enfoque multidisciplinario para cuidar tu salud física, mental y emocional.',
        'hero_img': 'https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'vitalmente centro medico, psicologia clinica roraima cdmx, centro de salud integral lomas, psicologo lomas cdmx',
        'servicios': [
            ('Evaluación & Consulta Psicológica', 'Diagnóstico preciso y plan terapéutico estructurado para alcanzar la estabilidad emocional.'),
            ('Terapia de Control de Ansiedad & Pánico', 'Técnicas clínicas probadas para detener ataques de pánico y angustia constante.'),
            ('Psicoterapia Familiar & Adolescentes', 'Intervención sistémica para restaurar la armonía y convivencia en el hogar.'),
            ('Atención Integral de Salud', 'Sinergia entre psicología y especialistas médicos para el bienestar del paciente.')
        ]
    }
}

def generate_html_content(item, config):
    nombre = item['nombre']
    ciudad = item['ciudad']
    tipo = item['tipo']
    tel = item['tel']
    resenas = item['resenas']
    rating = item['rating']
    demo_url = f"https://betto12680.github.io/borradoes-webs/{config['slug']}/"

    rating_val = str(rating) if rating and str(rating) != 'N/D' else '4.9'
    resenas_val = str(resenas) if resenas and str(resenas) != 'N/D' else '35'
    tel_clean = re.sub(r'\D', '', str(tel)) if tel and str(tel) != 'N/D' else '525555551234'
    if not tel_clean.startswith('52'):
        tel_clean = '52' + tel_clean

    nicho_schema = config['nicho']

    # FAQs por nicho
    if nicho_schema == 'Dentist':
        faqs = [
            ("¿Cómo puedo agendar mi primera cita de valoración?", "Puedes tocar el botón directo de WhatsApp para agendar tu fecha y hora disponible en menos de 2 minutos."),
            ("¿Aceptan urgencias u odontología el mismo día?", "Sí, atendemos emergencias por dolor agudo o fracturas dentales con prioridad según agenda."),
            ("¿Qué formas de pago aceptan?", "Aceptamos pago en efectivo, transferencias bancarias (SPEI) y tarjetas de crédito o débito."),
            ("¿Ofrecen facilidades de pago para tratamientos estéticos?", "Contamos con planes de financiamiento y facilidades en tratamientos de ortodoncia e implantes.")
        ]
    elif nicho_schema == 'PhysicalTherapy':
        faqs = [
            ("¿Necesito orden médica para iniciar fisioterapia?", "No es requisito indispensable. En tu primera valoración realizamos una exploración física completa."),
            ("¿Cuánto tiempo dura cada sesión de rehabilitación?", "Las sesiones duran habitualmente entre 50 y 60 minutos e incluyen terapia manual y aparatología."),
            ("¿Atienden a domicilio o solo en clínica?", "Brindamos atención en nuestras instalaciones equipadas y también contamos con servicio domiciliario según la zona."),
            ("¿Cuántas sesiones necesitaré para sentir mejoría?", "Muchos pacientes sienten alivio desde la primera sesión. El plan exacto se establece según el grado de tu lesión.")
        ]
    else:
        faqs = [
            ("¿Cómo agendar una consulta médica o psicológica?", "Haz clic en el botón de WhatsApp y nuestro equipo de recepción te confirmará los horarios disponibles."),
            ("¿Las consultas son presenciales u online?", "Ofrecemos atención presencial en nuestra clínica acondicionada y modalidad de consulta en línea previa cita."),
            ("¿Qué documentación o información debo llevar?", "Se recomienda acudir con historial médico previo o estudios recientes si cuentas con ellos."),
            ("¿Mantienen confidencialidad de expediente?", "Todos nuestros expedientes están protegidos bajo estricto secreto profesional y aviso de privacidad.")
        ]

    # JSON-LD Schema
    schema_json = f"""{{
  "@context": "https://schema.org",
  "@type": "{nicho_schema}",
  "name": "{nombre}",
  "url": "{demo_url}",
  "telephone": "+{tel_clean}",
  "address": {{
    "@type": "PostalAddress",
    "addressLocality": "{ciudad}",
    "addressCountry": "MX"
  }},
  "aggregateRating": {{
    "@type": "AggregateRating",
    "ratingValue": "{rating_val}",
    "reviewCount": "{resenas_val}"
  }},
  "priceRange": "$$"
}}"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{nombre} — {config['hero_title']}</title>
  <meta name="description" content="{config['hero_sub']}">
  <meta name="keywords" content="{config['keywords']}">
  <meta name="robots" content="index, follow">
  
  <!-- Open Graph / Facebook -->
  <meta property="og:type" content="website">
  <meta property="og:url" content="{demo_url}">
  <meta property="og:title" content="{nombre} — {config['hero_title']}">
  <meta property="og:description" content="{config['hero_sub']}">
  <meta property="og:image" content="{config['hero_img']}">
  <meta property="og:locale" content="es_MX">

  <!-- Twitter Card -->
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:title" content="{nombre}">
  <meta name="twitter:description" content="{config['hero_sub']}">
  <meta name="twitter:image" content="{config['hero_img']}">

  <!-- Fonts & Styles -->
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap" rel="stylesheet">
  
  <script type="application/ld+json">
  {schema_json}
  </script>

  <style>
    :root {{
      --primary: {config['primary']};
      --accent: {config['accent']};
      --bg: #0b0f19;
      --card-bg: #151c2c;
      --text: #f8fafc;
      --text-muted: #94a3b8;
      --border: #232d42;
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    html {{ scroll-behavior: smooth; }}
    body {{
      font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
      background-color: var(--bg);
      color: var(--text);
      line-height: 1.6;
      overflow-x: hidden;
    }}

    .container {{ max-width: 1140px; margin: 0 auto; padding: 0 24px; }}

    /* Navigation Header */
    nav {{
      position: fixed;
      top: 0; left: 0; right: 0;
      z-index: 1000;
      background: rgba(11, 15, 25, 0.85);
      backdrop-filter: blur(12px);
      -webkit-backdrop-filter: blur(12px);
      border-bottom: 1px solid var(--border);
      padding: 16px 0;
      transition: all 0.3s ease;
    }}
    .nav-inner {{ display: flex; justify-content: space-between; align-items: center; }}
    .brand-logo {{
      font-size: 1.35rem; font-weight: 800; color: #fff;
      text-decoration: none; display: flex; align-items: center; gap: 10px;
    }}
    .brand-logo span {{ color: var(--accent); }}

    .btn-header-wa {{
      background: #25d366; color: #fff; text-decoration: none;
      padding: 10px 20px; border-radius: 50px; font-weight: 700;
      font-size: 0.92rem; display: inline-flex; align-items: center; gap: 8px;
      box-shadow: 0 4px 14px rgba(37, 211, 102, 0.35);
      transition: transform 0.25s ease, box-shadow 0.25s ease;
    }}
    .btn-header-wa:hover {{
      transform: translateY(-2px);
      box-shadow: 0 6px 20px rgba(37, 211, 102, 0.5);
    }}

    /* Floating WhatsApp Button */
    .wa-float {{
      position: fixed; bottom: 28px; right: 28px; z-index: 999;
      background: #25d366; color: #fff; width: 62px; height: 62px;
      border-radius: 50%; display: flex; align-items: center; justify-content: center;
      box-shadow: 0 8px 24px rgba(37, 211, 102, 0.45);
      text-decoration: none; font-size: 1.8rem;
      transition: transform 0.3s ease;
    }}
    .wa-float:hover {{ transform: scale(1.1); }}
    .wa-float .pulse {{
      position: absolute; inset: 0; border-radius: 50%;
      background: #25d366; animation: wapulse 2.2s ease-out infinite; z-index: -1;
    }}
    @keyframes wapulse {{ 0% {{ transform: scale(1); opacity: 0.6; }} 100% {{ transform: scale(2); opacity: 0; }} }}

    /* Hero Section */
    .hero {{
      padding: 170px 0 100px;
      background: {config['bg_gradient']};
      position: relative; overflow: hidden; text-align: center;
    }}
    .hero::before {{
      content: ''; position: absolute; inset: 0;
      background: radial-gradient(circle at 50% 30%, rgba(255,255,255,0.08) 0%, transparent 60%);
      pointer-events: none;
    }}
    .hero-badge {{
      display: inline-flex; align-items: center; gap: 8px;
      background: rgba(255,255,255,0.1); backdrop-filter: blur(8px);
      border: 1px solid rgba(255,255,255,0.2);
      padding: 8px 20px; border-radius: 30px; font-weight: 600;
      font-size: 0.95rem; color: #fbbf24; margin-bottom: 24px;
    }}
    .hero h1 {{
      font-size: 3.1rem; font-weight: 800; max-width: 900px;
      margin: 0 auto 24px; line-height: 1.18; letter-spacing: -0.02em;
    }}
    .hero p.lead {{
      font-size: 1.2rem; color: rgba(255,255,255,0.85);
      max-width: 720px; margin: 0 auto 40px; font-weight: 400;
    }}
    .hero-actions {{ display: flex; justify-content: center; gap: 16px; flex-wrap: wrap; }}

    .btn-hero-primary {{
      background: #25d366; color: #fff; text-decoration: none;
      padding: 16px 36px; border-radius: 50px; font-weight: 800;
      font-size: 1.1rem; box-shadow: 0 10px 30px rgba(37, 211, 102, 0.4);
      transition: transform 0.25s ease, box-shadow 0.25s ease;
    }}
    .btn-hero-primary:hover {{ transform: translateY(-3px); box-shadow: 0 14px 36px rgba(37, 211, 102, 0.55); }}

    /* Image Banner Showcase */
    .hero-img-wrap {{
      margin-top: 60px; border-radius: 24px; overflow: hidden;
      border: 1px solid rgba(255,255,255,0.15);
      box-shadow: 0 25px 50px -12px rgba(0,0,0,0.6);
      max-height: 480px;
    }}
    .hero-img-wrap img {{ width: 100%; height: 100%; object-fit: cover; display: block; }}

    /* Services Grid */
    .section-title {{ text-align: center; margin-bottom: 50px; }}
    .section-title h2 {{ font-size: 2.3rem; font-weight: 800; margin-bottom: 12px; }}
    .section-title p {{ color: var(--text-muted); font-size: 1.1rem; max-width: 600px; margin: 0 auto; }}

    .services-grid {{
      display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 28px; margin-bottom: 90px;
    }}
    .service-card {{
      background: var(--card-bg); border: 1px solid var(--border);
      border-radius: 20px; padding: 32px; transition: all 0.3s ease;
      position: relative; overflow: hidden;
    }}
    .service-card::before {{
      content: ''; position: absolute; top: 0; left: 0; width: 100%; height: 4px;
      background: linear-gradient(90deg, var(--primary), var(--accent));
      opacity: 0; transition: opacity 0.3s ease;
    }}
    .service-card:hover {{
      transform: translateY(-6px); border-color: var(--accent);
      box-shadow: 0 20px 40px -15px rgba(0,0,0,0.5);
    }}
    .service-card:hover::before {{ opacity: 1; }}
    .service-icon {{
      width: 52px; height: 52px; background: rgba(59, 130, 246, 0.1);
      border-radius: 14px; display: flex; align-items: center; justify-content: center;
      font-size: 1.5rem; margin-bottom: 20px; color: var(--accent);
    }}
    .service-card h3 {{ font-size: 1.3rem; font-weight: 700; margin-bottom: 12px; color: #fff; }}
    .service-card p {{ color: var(--text-muted); font-size: 0.98rem; line-height: 1.6; }}

    /* Trust / Testimonials Banner */
    .trust-banner {{
      background: var(--card-bg); border: 1px solid var(--border);
      border-radius: 24px; padding: 48px; margin-bottom: 90px; text-align: center;
    }}
    .trust-rating-score {{ font-size: 3.5rem; font-weight: 800; color: #fbbf24; line-height: 1; margin-bottom: 8px; }}
    .trust-stars {{ color: #fbbf24; font-size: 1.5rem; margin-bottom: 12px; }}
    .trust-reviews-count {{ color: var(--text-muted); font-size: 1.05rem; margin-bottom: 28px; }}

    /* FAQs Section */
    .faqs-wrap {{ max-width: 800px; margin: 0 auto 90px; }}
    .faq-card {{
      background: var(--card-bg); border: 1px solid var(--border);
      border-radius: 14px; margin-bottom: 16px; overflow: hidden;
    }}
    .faq-card summary {{
      padding: 22px 28px; font-weight: 700; font-size: 1.08rem;
      cursor: pointer; list-style: none; display: flex;
      justify-content: space-between; align-items: center; color: #fff;
    }}
    .faq-card summary::-webkit-details-marker {{ display: none; }}
    .faq-card summary::after {{
      content: '+'; font-size: 1.5rem; color: var(--accent); transition: transform 0.25s ease;
    }}
    .faq-card[open] summary::after {{ transform: rotate(45deg); }}
    .faq-card p {{ padding: 0 28px 22px; color: var(--text-muted); font-size: 1rem; border-top: 1px solid rgba(255,255,255,0.05); margin-top: 8px; padding-top: 16px; }}

    /* Footer */
    footer {{
      border-top: 1px solid var(--border); padding: 50px 0;
      text-align: center; color: var(--text-muted); font-size: 0.95rem;
    }}
    footer a {{ color: var(--accent); text-decoration: none; font-weight: 600; }}

    @media (max-width: 768px) {{
      .hero h1 {{ font-size: 2.2rem; }}
      .hero p.lead {{ font-size: 1.05rem; }}
      .btn-hero-primary {{ padding: 14px 28px; font-size: 1rem; width: 100%; text-align: center; }}
    }}
  </style>
</head>
<body>

  <!-- Floating WhatsApp CTA -->
  <a href="https://wa.me/{tel_clean}?text=Hola%20equipo%20de%20{nombre},%20me%20gustar%C3%ADa%20solicitar%20informaci%C3%B3n%20y%20agendar%20una%20cita." class="wa-float" target="_blank" aria-label="Contactar por WhatsApp">
    <div class="pulse"></div>
    💬
  </a>

  <!-- Header Navigation -->
  <nav>
    <div class="container nav-inner">
      <a href="#" class="brand-logo">
        {nombre}
      </a>
      <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20deseo%20agendar%20cita." class="btn-header-wa" target="_blank">
        💬 Agendar por WhatsApp
      </a>
    </div>
  </nav>

  <!-- Hero Showcase Section -->
  <header class="hero">
    <div class="container">
      <div class="hero-badge">
        ⭐ {rating_val} Calificación en Google Maps ({resenas_val} opiniones reales)
      </div>
      <h1>{config['hero_title']}</h1>
      <p class="lead">{config['hero_sub']}</p>
      <div class="hero-actions">
        <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20quiero%20m%C3%A1s%20informaci%C3%B3n" class="btn-hero-primary" target="_blank">
          👉 Agendar Mi Cita Directa por WhatsApp
        </a>
      </div>
      
      <div class="hero-img-wrap">
        <img src="{config['hero_img']}" alt="{nombre} - {config['hero_title']}" loading="lazy" width="1200" height="480">
      </div>
    </div>
  </header>

  <!-- Main Services Section -->
  <section class="container" style="padding-top: 90px;">
    <div class="section-title">
      <h2>Nuestros Servicios Especializados</h2>
      <p>Brindamos atención personalizada con los más altos estándares de calidad en {ciudad}.</p>
    </div>

    <div class="services-grid">
"""
    icons = ['✨', '💎', '🛡️', '⚡']
    for idx_s, (s_title, s_desc) in enumerate(config['servicios']):
        ico = icons[idx_s % len(icons)]
        html += f"""
      <div class="service-card">
        <div class="service-icon">{ico}</div>
        <h3>{s_title}</h3>
        <p>{s_desc}</p>
      </div>"""

    html += f"""
    </div>
  </section>

  <!-- Trust & Google Rating Showcase Banner -->
  <section class="container">
    <div class="trust-banner">
      <div class="trust-rating-score">{rating_val}</div>
      <div class="trust-stars">★★★★★</div>
      <div class="trust-reviews-count">Reseñas de pacientes y clientes verificados en Google Maps ({resenas_val} opiniones)</div>
      <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20me%20gustar%C3%ADa%20recibir%20atenci%C3%B3n" class="btn-header-wa" style="padding: 12px 30px; font-size: 1.05rem;" target="_blank">
        💬 Contactar por WhatsApp
      </a>
    </div>
  </section>

  <!-- FAQs Section -->
  <section class="container">
    <div class="section-title">
      <h2>Preguntas Frecuentes</h2>
      <p>Resolvemos tus dudas principales para que inicies tu tratamiento con total tranquilidad.</p>
    </div>

    <div class="faqs-wrap">
"""
    for q_title, q_ans in faqs:
        html += f"""
      <details class="faq-card">
        <summary>{q_title}</summary>
        <p>{q_ans}</p>
      </details>"""

    html += f"""
    </div>
  </section>

  <!-- Footer -->
  <footer>
    <div class="container">
      <p>&copy; {datetime.now().year} {nombre} — {ciudad}. Propuesta de diseño web por <a href="https://wa.me/573104816153" target="_blank">Edilberto Sarmiento</a>.</p>
    </div>
  </footer>

</body>
</html>
"""
    return html

def build_all_cdmx():
    print("=== EJECUTANDO PASO 2: CONSTRUCCIÓN Y DESPLIEGUE WEBS CDMX ===")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb['EMPRESAS MAPS']
    rows = list(sheet.iter_rows(values_only=False))

    header = [cell.value for cell in rows[0]]
    built_sites = []

    for idx, r in enumerate(rows[1:], start=2):
        vals = [cell.value for cell in r]
        if not any(vals): continue
        nombre = str(vals[0]).strip()
        ciudad = str(vals[1]).strip()
        tipo = str(vals[2]).strip()
        correo = str(vals[3]).strip() if vals[3] else ''
        tel = str(vals[4]).strip() if vals[4] else ''
        maps_link = str(vals[5]).strip() if vals[5] else ''
        resenas = str(vals[6]).strip() if vals[6] else ''
        rating = str(vals[7]).strip() if vals[7] else ''
        estado = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ''

        if 'PENDIENTE_BORRADOR' in estado or 'Prospecto' in estado or 'PENDIENTE' in estado:
            # Buscar configuración de marca o usar fallback dinámico
            config = BRANDS_CONFIG.get(nombre)
            if not config:
                slug_val = slugify(nombre)
                config = {
                    'slug': slug_val,
                    'nicho': 'Dentist' if 'dent' in tipo.lower() or 'odont' in tipo.lower() else ('PhysicalTherapy' if 'fisio' in tipo.lower() or 'rehab' in tipo.lower() else 'MedicalClinic'),
                    'primary': '#0284c7',
                    'accent': '#38bdf8',
                    'bg_gradient': 'linear-gradient(135deg, #0c4a6e 0%, #0284c7 100%)',
                    'hero_title': f'Atención Profesional y Especializada en {ciudad}',
                    'hero_sub': f'Servicios de excelencia y diagnóstico de primer nivel en {ciudad}. Tu bienestar es nuestra prioridad.',
                    'hero_img': 'https://images.unsplash.com/photo-1629909613654-28e377c37b09?auto=format&fit=crop&w=1200&q=80',
                    'keywords': f'{slug_val}, clinica cdmx, especialidades cdmx',
                    'servicios': [
                        ('Atención Personalizada', 'Evaluación detallada para atender tus necesidades de forma ágil y profesional.'),
                        ('Tecnología de Vanguardia', 'Equipos modernos para diagnósticos y procedimientos seguros.'),
                        ('Especialistas Certificados', 'Personal con alta preparación médica y enfoque humano.'),
                        ('Instalaciones Cómodas', 'Ambiente limpio, confortable y seguro en la Ciudad de México.')
                    ]
                }

            item = {
                'nombre': nombre,
                'ciudad': ciudad,
                'tipo': tipo,
                'tel': tel,
                'resenas': resenas,
                'rating': rating
            }

            slug = config['slug']
            site_dir = os.path.join(BASE_DIR, slug)
            os.makedirs(site_dir, exist_ok=True)

            html_code = generate_html_content(item, config)
            index_path = os.path.join(site_dir, 'index.html')
            with open(index_path, 'w', encoding='utf-8') as f:
                f.write(html_code)

            demo_url = f"https://betto12680.github.io/borradoes-webs/{slug}/"
            nuevo_estado = f"WEB_BORRADOR_LISTA | LINK: {demo_url}"
            sheet.cell(row=idx, column=10, value=nuevo_estado)

            built_sites.append((nombre, slug, demo_url))
            print(f"✔ Web borrador creada para [{nombre}] -> {demo_url}")

    if built_sites:
        wb.save(EXCEL_PATH)
        print(f"\nSe crearon {len(built_sites)} sitios web borrador en disco. Desplegando en GitHub...")

        try:
            subprocess.run(["git", "add", "."], cwd=BASE_DIR, check=True)
            subprocess.run(["git", "commit", "-m", f"Despliegue Paso 2: {len(built_sites)} webs borrador CDMX en GitHub Pages"], cwd=BASE_DIR, check=True)
            subprocess.run(["git", "push"], cwd=BASE_DIR, check=True)
            print("✔ Despliegue en GitHub Pages completado exitosamente.")
        except Exception as e:
            print(f"Error en Git push: {e}")
    else:
        print("No se encontraron empresas pendientes para generar web hoy.")

if __name__ == "__main__":
    build_all_cdmx()
