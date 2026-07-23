#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Paso 2: Generación de Briefs, Páginas Web Borrador V2 y Despliegue a GitHub Pages (Guadalajara)
"""

import os
import re
import glob
import subprocess
import openpyxl
from datetime import datetime

BASE_DIR = '/Users/beto/Library/CloudStorage/OneDrive-Personal(2)/Freelancer/clientes/web'
EXCEL_PATH = os.path.join(BASE_DIR, 'Clientes.xlsx')

def slugify(text):
    text = text.lower().strip()
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n', 'ü': 'u'}
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text.strip('-')

PROMPT_MAESTRO_TEMPLATE = """## 🤖 PROMPT MAESTRO PARA GENERACIÓN DE CÓDIGO (PARA LA IA)

**ROLES Y DIRECTRICES:**
Eres un Desarrollador Frontend Senior experto en UX/UI y diseño de alta conversión. Tu objetivo es tomar la información contenida en este brief y generar un archivo `index.html` altamente dinámico, pulido y profesional, basado en el estándar estético de `odontosaludp.com`.

**REGLAS ESTRICTAS DE DISEÑO:**
1. **NO USAR frameworks pesados**: Solo HTML5, CSS3 nativo, y JavaScript vanilla.
2. **CERO colores agresivos**: NUNCA usar rojo, negro absoluto o colores oscuros pesados. Utilizar EXCLUSIVAMENTE la paleta de colores definida en este brief (azules, turquesas, blancos, fondos médicos limpios).
3. **CERO Emojis**: NUNCA utilizar emojis en los textos ni botones. Utilizar viñetas animadas (pulse-bullet-dot) o iconos vectoriales SVG limpios.
4. **Secciones Obligatorias**:
   - Header sticky con efecto Glassmorphism (blur) y botón de contacto.
   - Hero Section dividido (Split layout) con imagen de alta calidad, copy persuasivo, y badge de calificaciones.
   - Barra de Estadísticas Animada (Contadores).
   - Sección de Servicios en formato Tarjetas con viñetas animadas de pulso (`pulse-bullet-dot`).
   - Carrusel/Slider de Instalaciones o Antes/Después (implementado nativamente en JS).
   - Sección de Preguntas Frecuentes (Acordeón interactivo).
   - Botón Flotante de WhatsApp con animación de "latido/pulso" y vector SVG.
5. **Animaciones**: Implementar `IntersectionObserver` para revelar elementos suavemente al hacer scroll (`fade-in-up`).
6. **Responsividad**: Perfecto en dispositivos móviles (flexbox/grid).
7. **Contenido**: Insertar exactamente los textos de este brief. Añadir sutilmente al footer: "*Este es un diseño borrador de propuesta.*"

**OUTPUT ESPERADO:**
Genera el código completo en un solo bloque HTML que integre todo el CSS y JS necesario para que funcione inmediatamente al abrirlo en el navegador. Reemplaza el `index.html` en esta carpeta con este nuevo código.
"""

def generate_brief(nombre, ciudad, tipo, tel, resenas, rating, slug):
    tipo_lower = tipo.lower()
    
    if 'odont' in tipo_lower or 'dent' in tipo_lower:
        paleta = """
- **Primario**: #0284c7 (Azul Médico)
- **Secundario**: #38bdf8 (Cian Suave)
- **Fondo Oscuro (Hero)**: #0c4a6e (Azul Profundo, NUNCA negro)
- **Acentos**: Blanco puro y dorado suave para estrellas de reseñas."""
        
        imagenes = """
- **Hero Image**: https://images.unsplash.com/photo-1606811841689-23dfddce3e95?auto=format&fit=crop&w=1200&q=80 (Clínica limpia y moderna)
- **Carrusel 1**: https://images.unsplash.com/photo-1598256989800-fe5f95da9787?auto=format&fit=crop&w=1200&q=80
- **Carrusel 2**: https://images.unsplash.com/photo-1629909613654-28e377c37b09?auto=format&fit=crop&w=1200&q=80
- **Carrusel 3**: https://images.unsplash.com/photo-1588776814546-1ffcf47267a5?auto=format&fit=crop&w=1200&q=80"""
        
        textos = f"""*(Texto borrador inventado)*
- **Hero Titular**: Odontología Avanzada y Diseño de Sonrisa en {ciudad}
- **Hero Subtítulo**: Especialistas certificados enfocados en devolverte la confianza al sonreír, con tecnología digital 3D y tratamientos sin dolor.
- **Servicios Principales**: 
  1. Diseño de Sonrisa en 3D
  2. Ortodoncia Invisible
  3. Implantes Dentales de Carga Inmediata
  4. Blanqueamiento Láser Clínico"""

    elif 'fisio' in tipo_lower or 'rehab' in tipo_lower or 'ortoped' in tipo_lower:
        paleta = """
- **Primario**: #0d9488 (Turquesa Salud)
- **Secundario**: #14b8a6 (Teal Brillante)
- **Fondo Oscuro (Hero)**: #115e59 (Turquesa Profundo, NUNCA negro)
- **Acentos**: Blanco, gris perla y toques de verde vitalidad."""
        
        imagenes = """
- **Hero Image**: https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80 (Fisioterapia, terapia manual)
- **Carrusel 1**: https://images.unsplash.com/photo-1519823551278-64ac92734fb1?auto=format&fit=crop&w=1200&q=80
- **Carrusel 2**: https://images.unsplash.com/photo-1534438327276-14e5300c3a48?auto=format&fit=crop&w=1200&q=80
- **Carrusel 3**: https://images.unsplash.com/photo-1544367567-0f2fcb009e0b?auto=format&fit=crop&w=1200&q=80"""
        
        textos = f"""*(Texto borrador inventado)*
- **Hero Titular**: Recupera tu Movilidad y Dile Adiós al Dolor en {ciudad}
- **Hero Subtítulo**: Fisioterapia deportiva, neurológica y traumatológica guiada por especialistas. Recuperación acelerada con la mejor tecnología.
- **Servicios Principales**: 
  1. Terapia Manual y Liberación Miofascial
  2. Rehabilitación Deportiva
  3. Fisioterapia Post-Operatoria
  4. Punción Seca y Electrólisis"""

    else:
        paleta = """
- **Primario**: #2563eb (Azul Clínico Real)
- **Secundario**: #60a5fa (Azul Claro)
- **Fondo Oscuro (Hero)**: #1e3a8a (Azul Marino, NUNCA negro)
- **Acentos**: Blanco, gris claro para fondos de tarjetas."""
        
        imagenes = """
- **Hero Image**: https://images.unsplash.com/photo-1516549655169-df83a0774514?auto=format&fit=crop&w=1200&q=80 (Ambiente clínico)
- **Carrusel 1**: https://images.unsplash.com/photo-1551076805-e18690c5e53b?auto=format&fit=crop&w=1200&q=80
- **Carrusel 2**: https://images.unsplash.com/photo-1581056771107-24ca5f033842?auto=format&fit=crop&w=1200&q=80
- **Carrusel 3**: https://images.unsplash.com/photo-1579684385127-1ef15d508118?auto=format&fit=crop&w=1200&q=80"""
        
        textos = f"""*(Texto borrador inventado)*
- **Hero Titular**: Atención Médica de Excelencia en {ciudad}
- **Hero Subtítulo**: Un equipo multidisciplinario comprometido con tu salud integral. Instalaciones de primer nivel y trato humano cálido.
- **Servicios Principales**: 
  1. Consulta Médica Especializada
  2. Diagnóstico y Laboratorio Clínico
  3. Checkups Integrales Preventivos
  4. Especialidades Quirúrgicas"""

    rating_val = str(rating) if rating and str(rating) != 'N/D' else '4.9'
    resenas_val = str(resenas) if resenas and str(resenas) != 'N/D' else '50'
    tel_clean = re.sub(r'\D', '', str(tel)) if tel and str(tel) != 'N/D' else '523336151234'
    if not tel_clean.startswith('52'): tel_clean = '52' + tel_clean

    brief = f"""# Brief de Desarrollo Web: {nombre}

## 1. Información del Cliente
- **Nombre**: {nombre}
- **Nicho**: {tipo}
- **Ciudad**: {ciudad}
- **Teléfono de Contacto**: {tel} (WhatsApp: {tel_clean})
- **Calificación en Google Maps**: {rating_val} Estrellas
- **Cantidad de Reseñas**: {resenas_val} reseñas

## 2. Paleta de Colores Autorizada
**IMPORTANTE: NUNCA usar rojo, ni colores relacionados con la sangre, la violencia o el peligro, ni fondos negros puros. Todo debe transmitir limpieza, paz, clínica, salud y profesionalismo.**
{paleta}

## 3. Banco de Imágenes de Referencia (Unsplash)
{imagenes}

## 4. Estructura y Textos del Borrador
{textos}

---

{PROMPT_MAESTRO_TEMPLATE}
"""
    return brief, tel_clean, rating_val, resenas_val

def generate_html(nombre, ciudad, tipo, tel_clean, rating_val, resenas_val, slug):
    tipo_lower = tipo.lower()
    
    if 'odont' in tipo_lower or 'dent' in tipo_lower:
        primary = "#0284c7"
        secondary = "#38bdf8"
        dark_bg = "#0c4a6e"
        hero_title = f"Odontología Avanzada y Diseño de Sonrisa en {ciudad}"
        hero_sub = "Especialistas certificados enfocados en devolverte la confianza al sonreír, con tecnología digital 3D y tratamientos sin dolor."
        hero_img = "https://images.unsplash.com/photo-1606811841689-23dfddce3e95?auto=format&fit=crop&w=1200&q=80"
        carousel_imgs = [
            "https://images.unsplash.com/photo-1598256989800-fe5f95da9787?auto=format&fit=crop&w=1200&q=80",
            "https://images.unsplash.com/photo-1629909613654-28e377c37b09?auto=format&fit=crop&w=1200&q=80",
            "https://images.unsplash.com/photo-1588776814546-1ffcf47267a5?auto=format&fit=crop&w=1200&q=80"
        ]
        servicios = [
            ("Diseño de Sonrisa en 3D", "Diseño digital 3D de tu sonrisa antes de comenzar el tratamiento. Planificación con precisión milimétrica."),
            ("Ortodoncia Invisible", "Alineadores transparentes prácticamente imperceptibles. Corrige la posición de tus dientes sin brackets metálicos."),
            ("Implantes Dentales de Carga Inmediata", "Implantes de titanio biocompatible con corona inmediata. Recupera la funcionalidad y estética en una sesión."),
            ("Blanqueamiento Láser Clínico", "Aclarado dental profesional con luz LED intensiva. Resultados visibles en una sola sesión sin sensibilidad.")
        ]
    elif 'fisio' in tipo_lower or 'rehab' in tipo_lower or 'ortoped' in tipo_lower:
        primary = "#0d9488"
        secondary = "#14b8a6"
        dark_bg = "#115e59"
        hero_title = f"Recupera tu Movilidad y Dile Adiós al Dolor en {ciudad}"
        hero_sub = "Fisioterapia deportiva, neurológica y traumatológica guiada por especialistas. Recuperación acelerada con la mejor tecnología."
        hero_img = "https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80"
        carousel_imgs = [
            "https://images.unsplash.com/photo-1519823551278-64ac92734fb1?auto=format&fit=crop&w=1200&q=80",
            "https://images.unsplash.com/photo-1534438327276-14e5300c3a48?auto=format&fit=crop&w=1200&q=80",
            "https://images.unsplash.com/photo-1544367567-0f2fcb009e0b?auto=format&fit=crop&w=1200&q=80"
        ]
        servicios = [
            ("Terapia Manual y Liberación Miofascial", "Tratamiento enfocado en aliviar la tensión profunda en músculos y articulaciones acumulada por estrés o lesiones."),
            ("Rehabilitación Deportiva", "Optimización muscular y protocolos personalizados para retornar a tus entrenamientos con total seguridad."),
            ("Fisioterapia Post-Operatoria", "Acompañamiento especializado para reacondicionar tejidos y acelerar la cicatrización tras intervenciones quirúrgicas."),
            ("Punción Seca y Electrólisis", "Técnicas avanzadas para desactivar puntos gatillo musculares y reducir la inflamación de forma rápida.")
        ]
    else:
        primary = "#2563eb"
        secondary = "#60a5fa"
        dark_bg = "#1e3a8a"
        hero_title = f"Atención Médica de Excelencia en {ciudad}"
        hero_sub = "Un equipo multidisciplinario comprometido con tu salud integral. Instalaciones de primer nivel y trato humano cálido."
        hero_img = "https://images.unsplash.com/photo-1516549655169-df83a0774514?auto=format&fit=crop&w=1200&q=80"
        carousel_imgs = [
            "https://images.unsplash.com/photo-1551076805-e18690c5e53b?auto=format&fit=crop&w=1200&q=80",
            "https://images.unsplash.com/photo-1581056771107-24ca5f033842?auto=format&fit=crop&w=1200&q=80",
            "https://images.unsplash.com/photo-1579684385127-1ef15d508118?auto=format&fit=crop&w=1200&q=80"
        ]
        servicios = [
            ("Consulta Médica Especializada", "Valoración clínica detallada y personalizada para un diagnóstico certero y prevención oportuna."),
            ("Diagnóstico y Laboratorio Clínico", "Estudios médicos con tecnología avanzada para obtener resultados rápidos y confiables."),
            ("Checkups Integrales Preventivos", "Evaluación integral de salud diseñada para monitorear tu estado general y prevenir riesgos."),
            ("Especialidades Quirúrgicas", "Atención multidisciplinaria para procedimientos médicos con altos estándares de bioseguridad.")
        ]

    # SVGs for non-emoji elements
    svg_wa_header = '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="display:inline-block; vertical-align:middle; margin-right:6px;"><path d="M7.9 20A9 9 0 1 0 4 16.1L2 22Z"/></svg>'
    svg_wa_float = '<svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M7.9 20A9 9 0 1 0 4 16.1L2 22Z"/></svg>'
    svg_star = '<svg width="16" height="16" viewBox="0 0 24 24" fill="#fbbf24" stroke="#fbbf24" stroke-width="2" style="display:inline-block; vertical-align:middle; margin-right:4px;"><polygon points="12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2"/></svg>'

    servicios_html = ""
    for s_title, s_desc in servicios:
        servicios_html += f"""
        <div class="srv-card fade-in-up">
            <div class="pulse-bullet-dot"><span class="dot"></span></div>
            <h3>{s_title}</h3>
            <p>{s_desc}</p>
            <div class="tags"><span class="tag">Especializado</span><span class="tag">Seguro</span></div>
            <a href="https://wa.me/{tel_clean}?text=Hola%2C%20me%20interesa%20el%20servicio%20de%20{s_title}" class="btn-outline">Consultar por WhatsApp</a>
        </div>"""

    carousel_html = ""
    for idx, img in enumerate(carousel_imgs):
        active = 'active' if idx == 0 else ''
        carousel_html += f"""<div class="carousel-slide {active}"><img src="{img}" alt="Instalaciones {nombre}"></div>"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{nombre} | {tipo}</title>
    <meta name="description" content="{hero_sub}">
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        :root {{
            --primary: {primary};
            --secondary: {secondary};
            --dark-bg: {dark_bg};
            --bg: #070f1a;
            --card-bg: #0e1828;
            --border: #1a2a40;
            --text: #f8fafc;
            --text-muted: #94a3b8;
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        html {{ scroll-behavior: smooth; }}
        body {{ font-family: 'Plus Jakarta Sans', sans-serif; background-color: var(--bg); color: var(--text); overflow-x: hidden; }}
        
        .container {{ max-width: 1200px; margin: 0 auto; padding: 0 24px; }}
        
        /* Navbar Glassmorphism */
        nav {{ position: fixed; top: 0; width: 100%; z-index: 1000; background: rgba(7, 15, 26, 0.75); backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px); border-bottom: 1px solid rgba(255,255,255,0.05); padding: 16px 0; transition: all 0.3s ease; }}
        .nav-content {{ display: flex; justify-content: space-between; align-items: center; }}
        .brand {{ font-size: 1.4rem; font-weight: 800; color: #fff; text-decoration: none; }}
        .brand span {{ color: var(--secondary); }}
        .btn-wa-header {{ background: var(--primary); color: #fff; padding: 10px 22px; border-radius: 50px; text-decoration: none; font-weight: 700; display: inline-flex; align-items: center; transition: all 0.3s ease; box-shadow: 0 4px 15px rgba(0,0,0,0.2); }}
        .btn-wa-header:hover {{ background: var(--secondary); transform: translateY(-2px); }}

        /* Hero Split Layout */
        .hero {{ padding: 140px 0 80px; background: radial-gradient(circle at top left, var(--dark-bg) 0%, var(--bg) 70%); }}
        .hero-grid {{ display: grid; grid-template-columns: 1.1fr 0.9fr; gap: 40px; align-items: center; }}
        .hero-badge {{ display: inline-flex; background: rgba(255,255,255,0.08); padding: 8px 16px; border-radius: 20px; font-weight: 700; color: var(--secondary); margin-bottom: 24px; font-size: 0.88rem; border: 1px solid rgba(255,255,255,0.1); text-transform: uppercase; letter-spacing: 0.5px; }}
        .hero h1 {{ font-size: 3.2rem; font-weight: 800; line-height: 1.15; margin-bottom: 24px; }}
        .hero h1 span {{ color: var(--secondary); }}
        .hero p {{ font-size: 1.1rem; color: var(--text-muted); margin-bottom: 32px; line-height: 1.65; }}
        .hero-cta {{ display: flex; gap: 16px; align-items: center; }}
        .btn-primary {{ background: #25d366; color: #fff; padding: 16px 32px; border-radius: 50px; text-decoration: none; font-weight: 800; font-size: 1.05rem; display: inline-flex; align-items: center; transition: all 0.3s ease; box-shadow: 0 8px 24px rgba(37, 211, 102, 0.3); }}
        .btn-primary:hover {{ transform: translateY(-3px); box-shadow: 0 12px 32px rgba(37, 211, 102, 0.5); }}
        .hero-rating {{ display: flex; align-items: center; gap: 12px; margin-top: 32px; font-size: 0.9rem; color: var(--text-muted); }}
        .hero-rating strong {{ color: #fbbf24; font-size: 1.1rem; }}
        
        .hero-image-wrap {{ position: relative; border-radius: 24px; overflow: hidden; box-shadow: 0 24px 48px rgba(0,0,0,0.4); border: 1px solid rgba(255,255,255,0.1); }}
        .hero-image-wrap img {{ width: 100%; height: 480px; object-fit: cover; display: block; }}
        
        /* Stats Bar */
        .stats {{ background: var(--card-bg); padding: 40px 0; border-top: 1px solid var(--border); border-bottom: 1px solid var(--border); }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); text-align: center; gap: 20px; }}
        .stat-item h3 {{ font-size: 2.5rem; color: var(--secondary); font-weight: 800; margin-bottom: 8px; }}
        .stat-item p {{ font-size: 0.9rem; color: var(--text-muted); font-weight: 600; text-transform: uppercase; letter-spacing: 1px; }}

        /* Viñetas Animadas (Punticos con Pulso) */
        .pulse-bullet-dot {{ display: inline-flex; align-items: center; justify-content: center; width: 24px; height: 24px; margin-bottom: 16px; }}
        .pulse-bullet-dot .dot {{ width: 10px; height: 10px; border-radius: 50%; background-color: var(--secondary); position: relative; box-shadow: 0 0 10px var(--secondary); }}
        .pulse-bullet-dot .dot::after {{ content: ''; position: absolute; top: -4px; left: -4px; right: -4px; bottom: -4px; border-radius: 50%; border: 2px solid var(--secondary); animation: pulse-ring 2s cubic-bezier(0.215, 0.61, 0.355, 1) infinite; }}
        @keyframes pulse-ring {{ 0% {{ transform: scale(0.5); opacity: 0.9; }} 100% {{ transform: scale(1.8); opacity: 0; }} }}

        /* Services Cards */
        .section-title {{ text-align: center; margin-bottom: 50px; }}
        .section-title h2 {{ font-size: 2.4rem; font-weight: 800; margin-bottom: 16px; }}
        .section-title p {{ color: var(--text-muted); max-width: 600px; margin: 0 auto; }}
        .services {{ padding: 100px 0; }}
        .srv-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(270px, 1fr)); gap: 24px; }}
        .srv-card {{ background: var(--card-bg); border: 1px solid var(--border); border-radius: 20px; padding: 32px; transition: all 0.3s ease; position: relative; overflow: hidden; }}
        .srv-card:hover {{ transform: translateY(-8px); border-color: var(--secondary); box-shadow: 0 20px 40px rgba(0,0,0,0.3); }}
        .srv-card h3 {{ font-size: 1.25rem; margin-bottom: 14px; font-weight: 700; }}
        .srv-card p {{ color: var(--text-muted); font-size: 0.95rem; margin-bottom: 24px; line-height: 1.6; }}
        .tags {{ display: flex; gap: 8px; margin-bottom: 24px; }}
        .tag {{ font-size: 0.75rem; background: rgba(255,255,255,0.05); padding: 4px 10px; border-radius: 20px; color: var(--secondary); border: 1px solid rgba(255,255,255,0.1); }}
        .btn-outline {{ display: block; text-align: center; border: 1px solid var(--border); padding: 12px; border-radius: 12px; color: #fff; text-decoration: none; font-weight: 600; font-size: 0.9rem; transition: all 0.3s; }}
        .btn-outline:hover {{ background: var(--secondary); border-color: var(--secondary); color: #000; }}

        /* Carousel JS Native */
        .carousel-sec {{ padding: 100px 0; background: var(--card-bg); border-top: 1px solid var(--border); border-bottom: 1px solid var(--border); }}
        .carousel-wrap {{ position: relative; max-width: 900px; margin: 0 auto; border-radius: 24px; overflow: hidden; box-shadow: 0 20px 50px rgba(0,0,0,0.5); border: 1px solid rgba(255,255,255,0.1); }}
        .carousel-slide {{ display: none; width: 100%; height: 480px; }}
        .carousel-slide.active {{ display: block; animation: fade 0.8s ease-in-out; }}
        .carousel-slide img {{ width: 100%; height: 100%; object-fit: cover; }}
        @keyframes fade {{ from {{opacity: 0.4}} to {{opacity: 1}} }}
        .car-btn {{ position: absolute; top: 50%; transform: translateY(-50%); background: rgba(0,0,0,0.6); color: white; border: 1px solid rgba(255,255,255,0.2); width: 48px; height: 48px; border-radius: 50%; cursor: pointer; display: flex; align-items: center; justify-content: center; font-size: 1.2rem; backdrop-filter: blur(4px); transition: all 0.3s; }}
        .car-btn:hover {{ background: var(--primary); }}
        .car-prev {{ left: 20px; }}
        .car-next {{ right: 20px; }}

        /* FAQs */
        .faqs {{ padding: 100px 0; }}
        .faq-wrap {{ max-width: 800px; margin: 0 auto; }}
        .faq-item {{ background: var(--card-bg); border: 1px solid var(--border); border-radius: 16px; margin-bottom: 16px; }}
        .faq-item summary {{ padding: 24px; font-size: 1.05rem; font-weight: 700; cursor: pointer; list-style: none; display: flex; justify-content: space-between; }}
        .faq-item summary::-webkit-details-marker {{ display: none; }}
        .faq-item summary::after {{ content: '+'; color: var(--secondary); font-size: 1.4rem; }}
        .faq-item[open] summary::after {{ content: '−'; }}
        .faq-item p {{ padding: 0 24px 24px; color: var(--text-muted); border-top: 1px solid rgba(255,255,255,0.05); margin-top: 8px; padding-top: 16px; }}

        /* Floating WhatsApp */
        .wa-float {{ position: fixed; bottom: 30px; right: 30px; background: #25d366; width: 60px; height: 60px; border-radius: 50%; display: flex; justify-content: center; align-items: center; color: white; text-decoration: none; box-shadow: 0 10px 24px rgba(37, 211, 102, 0.4); z-index: 999; transition: transform 0.3s ease; }}
        .wa-float:hover {{ transform: scale(1.1); }}
        .wa-float::before {{ content: ''; position: absolute; inset: 0; border-radius: 50%; border: 2px solid #25d366; animation: waPulse 2s infinite; }}
        @keyframes waPulse {{ 0% {{ transform: scale(1); opacity: 1; }} 100% {{ transform: scale(1.6); opacity: 0; }} }}

        /* Scroll Animations */
        .fade-in-up {{ opacity: 0; transform: translateY(30px); transition: opacity 0.8s ease, transform 0.8s ease; }}
        .fade-in-up.visible {{ opacity: 1; transform: translateY(0); }}
        
        footer {{ text-align: center; padding: 40px 0; border-top: 1px solid var(--border); color: var(--text-muted); font-size: 0.9rem; }}
        
        @media(max-width: 900px) {{
            .hero-grid {{ grid-template-columns: 1fr; text-align: center; }}
            .hero-cta {{ justify-content: center; }}
            .hero-rating {{ justify-content: center; }}
            .stats-grid {{ grid-template-columns: 1fr 1fr; }}
            .carousel-slide {{ height: 350px; }}
        }}
    </style>
</head>
<body>

    <!-- Nav -->
    <nav>
        <div class="container nav-content">
            <a href="#" class="brand">{nombre}</a>
            <a href="https://wa.me/{tel_clean}" class="btn-wa-header" target="_blank">{svg_wa_header} Contactar</a>
        </div>
    </nav>

    <!-- Hero -->
    <header class="hero">
        <div class="container hero-grid">
            <div class="hero-text fade-in-up">
                <div class="hero-badge">Atención Especializada en {ciudad}</div>
                <h1>{hero_title}</h1>
                <p>{hero_sub}</p>
                <div class="hero-cta">
                    <a href="https://wa.me/{tel_clean}?text=Hola,%20deseo%20agendar%20una%20cita" class="btn-primary" target="_blank">{svg_wa_header} Agendar Cita por WhatsApp</a>
                </div>
                <div class="hero-rating">
                    <span>{svg_star} <strong>{rating_val} / 5.0</strong></span>
                    <span>Basado en {resenas_val} reseñas verificadas</span>
                </div>
            </div>
            <div class="hero-image-wrap fade-in-up" style="transition-delay: 0.2s;">
                <img src="{hero_img}" alt="{nombre}">
            </div>
        </div>
    </header>

    <!-- Stats -->
    <section class="stats">
        <div class="container stats-grid">
            <div class="stat-item fade-in-up"><h3 class="counter" data-target="1200">0</h3><p>Pacientes Felices</p></div>
            <div class="stat-item fade-in-up"><h3 class="counter" data-target="12">0</h3><p>Años Exp.</p></div>
            <div class="stat-item fade-in-up"><h3>100%</h3><p>Garantía de Calidad</p></div>
            <div class="stat-item fade-in-up"><h3>{rating_val}</h3><p>Calificación Google</p></div>
        </div>
    </section>

    <!-- Services -->
    <section class="services container" id="servicios">
        <div class="section-title fade-in-up">
            <h2>Nuestros Servicios</h2>
            <p>Descubre los tratamientos especializados que {nombre} ofrece con la más alta calidad y atención humana.</p>
        </div>
        <div class="srv-grid">
            {servicios_html}
        </div>
    </section>

    <!-- Carousel -->
    <section class="carousel-sec">
        <div class="container">
            <div class="section-title fade-in-up">
                <h2>Instalaciones de Vanguardia</h2>
                <p>Espacios confortables e higienizados diseñados para tu absoluta bioseguridad.</p>
            </div>
            <div class="carousel-wrap fade-in-up">
                {carousel_html}
                <button class="car-btn car-prev" onclick="moveSlide(-1)" aria-label="Anterior">❮</button>
                <button class="car-btn car-next" onclick="moveSlide(1)" aria-label="Siguiente">❯</button>
            </div>
        </div>
    </section>

    <!-- FAQs -->
    <section class="faqs container">
        <div class="section-title fade-in-up">
            <h2>Preguntas Frecuentes</h2>
            <p>Resolvemos tus dudas para que inicies tu tratamiento con absoluta confianza.</p>
        </div>
        <div class="faq-wrap fade-in-up">
            <details class="faq-item">
                <summary>¿Cómo puedo agendar una consulta?</summary>
                <p>Solo debes hacer clic en el botón de WhatsApp y nuestro equipo te asignará la fecha y hora disponible de inmediato.</p>
            </details>
            <details class="faq-item">
                <summary>¿Qué métodos de pago aceptan?</summary>
                <p>Aceptamos efectivo, transferencias bancarias (SPEI) y tarjetas de crédito o débito.</p>
            </details>
            <details class="faq-item">
                <summary>¿Ofrecen garantía o valoración inicial?</summary>
                <p>Todos nuestros procedimientos cuentan con garantía de calidad y seguimiento médico continuo.</p>
            </details>
        </div>
    </section>

    <!-- Footer -->
    <footer>
        <div class="container">
            <p>&copy; {datetime.now().year} {nombre} — {ciudad}. Todos los derechos reservados. | <i>Este es un diseño borrador de propuesta.</i></p>
        </div>
    </footer>

    <!-- Floating WhatsApp Button -->
    <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20quiero%20informaci%C3%B3n" class="wa-float" target="_blank" aria-label="WhatsApp Directo">
        {svg_wa_float}
    </a>

    <!-- JS Logic -->
    <script>
        // Scroll Observer
        const observer = new IntersectionObserver((entries) => {{
            entries.forEach(entry => {{
                if (entry.isIntersecting) {{
                    entry.target.classList.add('visible');
                }}
            }});
        }}, {{ threshold: 0.1 }});
        document.querySelectorAll('.fade-in-up').forEach(el => observer.observe(el));

        // Stats Counter
        const counters = document.querySelectorAll('.counter');
        const speed = 200;
        const counterObserver = new IntersectionObserver((entries, obs) => {{
            entries.forEach(entry => {{
                if (entry.isIntersecting) {{
                    const counter = entry.target;
                    const target = +counter.getAttribute('data-target');
                    const updateCount = () => {{
                        const count = +counter.innerText;
                        const inc = target / speed;
                        if(count < target) {{
                            counter.innerText = Math.ceil(count + inc);
                            setTimeout(updateCount, 20);
                        }} else {{
                            counter.innerText = target + (target > 100 ? '+' : '');
                        }}
                    }};
                    updateCount();
                    obs.unobserve(counter);
                }}
            }});
        }}, {{ threshold: 0.5 }});
        counters.forEach(c => counterObserver.observe(c));

        // Carousel
        let slideIndex = 0;
        const slides = document.querySelectorAll('.carousel-slide');
        function moveSlide(n) {{
            slides[slideIndex].classList.remove('active');
            slideIndex = (slideIndex + n + slides.length) % slides.length;
            slides[slideIndex].classList.add('active');
        }}
        if (slides.length > 0) {{
            setInterval(() => moveSlide(1), 5000);
        }}
    </script>
</body>
</html>"""
    return html

def audit_html(html):
    issues = []
    lower_html = html.lower()
    
    # 1. Color check (Aggressive colors or black background)
    forbidden = ['#ff0000', 'color: red', 'color:red', 'background: red', 'background:red', 'background-color: #000000', 'background: #000000']
    for f in forbidden:
        if f in lower_html:
            issues.append(f"Forbidden color/string found: {f}")
            
    # 2. Component check
    if 'intersectionobserver' not in lower_html:
        issues.append("Missing IntersectionObserver for animations.")
    if 'moveslide' not in lower_html:
        issues.append("Missing native JS Carousel logic.")
    if 'backdrop-filter: blur' not in lower_html:
        issues.append("Missing Glassmorphism CSS.")
    if 'pulse-bullet-dot' not in lower_html:
        issues.append("Missing pulse-bullet-dot bullet animation.")

    return issues

def main():
    print("=== PIPELINE PASO 2: GENERACIÓN DE BRIEFS, WEBS Y DESPLIEGUE A GITHUB (GUADALAJARA) ===")
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: No se encontró {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb['EMPRESAS MAPS']
    rows = list(sheet.iter_rows(values_only=False))

    built_count = 0
    audit_failed = 0

    for idx, r in enumerate(rows[1:], start=2):
        vals = [cell.value for cell in r]
        if not any(vals): continue
        nombre = str(vals[0]).strip()
        ciudad = str(vals[1]).strip()
        tipo = str(vals[2]).strip()
        correo = str(vals[3]).strip() if vals[3] else ''
        tel = str(vals[4]).strip() if vals[4] else ''
        resenas = str(vals[6]).strip() if vals[6] else ''
        rating = str(vals[7]).strip() if vals[7] else ''
        estado = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ''

        # Process Guadalajara leads or PENDIENTE state
        if ('Guadalajara' in ciudad and ('PENDIENTE' in estado or 'WEB_BORRADOR_LISTA' not in estado)):
            slug = slugify(nombre)
            site_dir = os.path.join(BASE_DIR, slug)
            os.makedirs(site_dir, exist_ok=True)

            # 1. Generate Brief Maestro
            brief_content, tel_clean, rating_val, resenas_val = generate_brief(nombre, ciudad, tipo, tel, resenas, rating, slug)
            brief_path = os.path.join(site_dir, 'brief_maestro.md')
            with open(brief_path, 'w', encoding='utf-8') as f:
                f.write(brief_content)

            # 2. Generate HTML
            html_code = generate_html(nombre, ciudad, tipo, tel_clean, rating_val, resenas_val, slug)

            # 3. Audit HTML
            issues = audit_html(html_code)
            if issues:
                print(f"❌ Auditoría fallida para {nombre}: {issues}")
                audit_failed += 1
                continue

            # 4. Write index.html
            html_path = os.path.join(site_dir, 'index.html')
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_code)

            # 5. Update Excel
            demo_url = f"https://betto12680.github.io/borradoes-webs/{slug}/"
            nuevo_estado = f"WEB_BORRADOR_LISTA | LINK: {demo_url}"
            sheet.cell(row=idx, column=10, value=nuevo_estado)

            built_count += 1
            print(f"✔ Web Paso 2 Creada & Auditada: [{nombre}] -> {demo_url}")

    wb.save(EXCEL_PATH)
    print(f"\n--- RESUMEN PASO 2 ---")
    print(f"Sitios Creados y Auditados: {built_count}")
    print(f"Auditorías Fallidas: {audit_failed}")

    if built_count > 0:
        print("\nDesplegando en GitHub Pages...")
        try:
            subprocess.run(["git", "add", "."], cwd=BASE_DIR, check=True)
            subprocess.run(["git", "commit", "-m", f"Paso 2 Guadalajara: {built_count} sitios web creados, auditados y desplegados"], cwd=BASE_DIR, check=True)
            subprocess.run(["git", "push"], cwd=BASE_DIR, check=True)
            print("✔ Despliegue en GitHub Pages completado exitosamente.")
        except Exception as e:
            print(f"❌ Error en Git push: {e}")

if __name__ == "__main__":
    main()
