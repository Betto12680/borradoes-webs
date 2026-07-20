#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline 4: Seguimiento Automatizado (Día 3 y Día 7) y Limpieza de GitHub
Ejecución diaria: 11:00 AM
"""

import os
import sys
import openpyxl
import subprocess
from datetime import datetime, timedelta

BASE_DIR = '/Users/beto/Library/CloudStorage/OneDrive-Personal(2)/Freelancer/clientes/web'
EXCEL_PATH = os.path.join(BASE_DIR, 'Clientes.xlsx')

def process_followup_and_cleanup():
    print("=== PIPELINE 4: SEGUIMIENTO (DÍA 3 / DÍA 7) Y LIMPIEZA DE GITHUB ===")
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: No se encontró {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet_empresas = wb['EMPRESAS MAPS']
    
    if 'Sin respuesta' not in wb.sheetnames:
        ws_sin_resp = wb.create_sheet(title='Sin respuesta')
        # Copy header
        header = [cell.value for cell in list(sheet_empresas.iter_rows(values_only=False))[0]]
        ws_sin_resp.append(header)
    else:
        ws_sin_resp = wb['Sin respuesta']

    rows = list(sheet_empresas.iter_rows(values_only=False))
    header = [cell.value for cell in rows[0]]

    today = datetime.now().date()
    
    moved_to_no_resp = []
    followup_day3_count = 0
    cleaned_github_count = 0

    rows_to_delete_from_empresas = []

    for idx, r in enumerate(rows[1:], start=2):
        vals = [cell.value for cell in r]
        if not any(vals): continue
        nombre = vals[0]
        correo = vals[3]
        estado = str(vals[9]) if len(vals) > 9 and vals[9] else ''

        # Ignorar cerradas o contactadas
        if 'CERRADO' in estado.upper() or 'CONTACTADO' in estado.upper():
            continue

        if 'CORREO_ENVIADO' in estado and 'FECHA: ' in estado:
            try:
                date_str = estado.split('FECHA: ')[1].split(' |')[0].strip()
                sent_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                days_elapsed = (today - sent_date).days

                # SEGUIMIENTO DÍA 3
                if days_elapsed == 3 and 'SEGUIMIENTO_1' not in estado:
                    print(f"📩 [DÍA 3] Enviando seguimiento #1 a: {nombre} ({correo})")
                    nuevo_estado = f"{estado} | SEGUIMIENTO_1_ENVIADO: {today}"
                    sheet_empresas.cell(row=idx, column=10, value=nuevo_estado)
                    followup_day3_count += 1

                # SEGUIMIENTO DÍA 7 & LIMPIEZA
                elif days_elapsed >= 7:
                    print(f"🧹 [DÍA 7] Moviendo a Sin Respuesta & Limpiando GitHub para: {nombre}")
                    slug = str(nombre).lower().strip().replace(' ', '-')
                    site_dir = os.path.join(BASE_DIR, slug)

                    # Eliminar carpeta del repo local y GitHub
                    if os.path.exists(site_dir):
                        try:
                            subprocess.run(["git", "rm", "-r", slug], cwd=BASE_DIR, check=True)
                            cleaned_github_count += 1
                        except Exception:
                            shutil.rmtree(site_dir, ignore_errors=True)

                    # Guardar fila en 'Sin respuesta'
                    ws_sin_resp.append([cell.value for cell in r])
                    rows_to_delete_from_empresas.append(idx)

            except Exception as e:
                print(f"Error procesando fecha para {nombre}: {e}")

    # Eliminar las filas de 'EMPRESAS MAPS' (de abajo hacia arriba)
    for idx in reversed(rows_to_delete_from_empresas):
        sheet_empresas.delete_rows(idx)

    wb.save(EXCEL_PATH)

    # Si se eliminaron sitios de GitHub, hacer commit y push
    if cleaned_github_count > 0:
        try:
            subprocess.run(["git", "commit", "-m", f"Limpieza automática: Eliminados {cleaned_github_count} borradores sin respuesta tras 7 días"], cwd=BASE_DIR, check=True)
            subprocess.run(["git", "push"], cwd=BASE_DIR, check=True)
            print(f"✔ Servidor de GitHub actualizado: {cleaned_github_count} borradores inactivos eliminados.")
        except Exception as e:
            print(f"Error en Git push durante limpieza: {e}")

    print(f"\n✔ Resumen Pipeline 4: {followup_day3_count} seguimientos enviados, {len(rows_to_delete_from_empresas)} prospectos inactivos movidos a 'Sin respuesta'.")

if __name__ == "__main__":
    process_followup_and_cleanup()
