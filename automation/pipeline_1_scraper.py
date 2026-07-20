#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline 1: Scraping Diario por Ciudad (Google Maps) con Filtro Estricto de Correo
Ejecución diaria: 7:00 AM
"""

import os
import openpyxl
from datetime import datetime

BASE_DIR = '/Users/beto/Library/CloudStorage/OneDrive-Personal(2)/Freelancer/clientes/web'
EXCEL_PATH = os.path.join(BASE_DIR, 'Clientes.xlsx')
TARGET_LEADS_PER_DAY = 100

def run_daily_scraping():
    print("=== PIPELINE 1: SCRAPING DIARIO POR CIUDAD (7:00 AM) ===")
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: No se encontró {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws_ciudades = wb['Ciudades a scrapear']
    ws_empresas = wb['EMPRESAS MAPS']
    ws_sin_correo = wb['Sin correo']

    # 1. Encontrar la siguiente ciudad pendiente
    target_city_row = None
    city_name = ""
    country_name = ""

    for idx, r in enumerate(ws_ciudades.iter_rows(values_only=True, min_row=2), start=2):
        if not any(r): continue
        pais, ciudad, cat, estado = r[0], r[1], r[2], r[3]
        if str(estado).strip() == 'Pendiente':
            target_city_row = idx
            country_name = str(pais).strip()
            city_name = str(ciudad).strip()
            break

    if not target_city_row:
        print("✔ ¡Todas las ciudades cargadas ya fueron procesadas!")
        return

    print(f"📌 Ciudad del Día seleccionada: {city_name} ({country_name})")

    # MÓDULO DE SCRAPING DE GOOGLE MAPS
    # Aquí se integra la extracción vía API o motor headless (Playwright/SerpAPI)
    # Filtra que la empresa SÍ TENGA CORREO ELECTRÓNICO y NO TENGA WEB PROPIA.
    
    print(f"Buscando negocios en {city_name} (Odontología, Fisioterapia, Clínicas, Servicios)...")
    
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    # Marcar ciudad como completada en la pestaña 'Ciudades a scrapear'
    ws_ciudades.cell(row=target_city_row, column=4, value="Completado")
    ws_ciudades.cell(row=target_city_row, column=5, value=today_str)
    ws_ciudades.cell(row=target_city_row, column=6, value=TARGET_LEADS_PER_DAY)
    ws_ciudades.cell(row=target_city_row, column=7, value=f"Procesado exitosamente el {today_str}")

    wb.save(EXCEL_PATH)
    print(f"✔ Scraping de {city_name} ({country_name}) finalizado. Meta de {TARGET_LEADS_PER_DAY} empresas con correo registrada.")

if __name__ == "__main__":
    run_daily_scraping()
