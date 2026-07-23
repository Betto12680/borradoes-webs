#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline 3: Envíos Personalizados por Correo Electrónico en Bloques Anti-Spam (API Mailrelay Oficial)
"""

import os
import sys
import time
import json
import urllib.request
import urllib.error
import openpyxl
from datetime import datetime

BASE_DIR = '/Users/beto/Library/CloudStorage/OneDrive-Personal(2)/Freelancer/clientes/web'
EXCEL_PATH = os.path.join(BASE_DIR, 'Clientes.xlsx')

# Configuración de Mailrelay de la marca (Trabajando con IA)
MAILRELAY_HOST = "trabajandoconia.ipzmarketing.com"
MAILRELAY_API_KEY = "Pk84z15yBRQYsMhj2mug7_af6Anb_gzsS28_qe_r"
SENDER_EMAIL = "contacto@trabajandoconia.com"
SENDER_NAME = "Edilberto Sarmiento"

BATCH_SIZE = 20  # Bloques de 20 correos
PAUSE_BETWEEN_BATCHES_SEC = 180  # 3 minutos de pausa entre lotes anti-spam
PAUSE_BETWEEN_EMAILS_SEC = 3     # 3 segundos entre cada correo dentro del bloque

def get_email_template(niche, nombre, demo_url):
    tipo_lower = str(niche).lower()
    nombre_lower = str(nombre).lower()
    
    if 'odont' in tipo_lower or 'denti' in tipo_lower or 'sonris' in nombre_lower:
        asunto = f"¿Sabes cuántos pacientes buscan odontólogo en Google y no te encuentran, {nombre}?"
        body_text = f"""Hola, equipo de {nombre}, ¿cómo están?

Estaba revisando perfiles de odontología en Google Maps y vi la excelente reputación y atención que tienen. Sin embargo, noté que cuando un nuevo paciente busca sus servicios en internet, no encuentra una página web oficial donde ver sus tratamientos o agendar una cita.

Hoy en día, la mayoría de personas prefieren revisar la web de un consultorio desde su celular antes de agendar su valoración.

Por eso, me tomé la libertad de diseñarles una primera propuesta de página web, completa y adaptada a su consultorio. Pueden verla funcionando en vivo desde su celular en este enlace:

👉 {demo_url}

📌 Importante sobre este borrador: Esto es tan solo una primera propuesta inicial de muestra (un borrador visual). Si les gusta la idea, la trabajamos juntos y la personalizamos al 100% con sus fotos reales, su logo oficial, sus servicios/tarifas exactas y todos los cambios que desees hacer antes de hacer el lanzamiento oficial en tu propio dominio.

Échenle una mirada sin ningún compromiso. Si les parece interesante, me avisan y con gusto conversamos.

Un saludo,

Edilberto Sarmiento
Diseñador Web & Estratega Digital
📱 WhatsApp: +57 310 481 6153"""

    elif 'glamp' in tipo_lower or 'hotel' in tipo_lower or 'hospedaj' in tipo_lower or 'cabañ' in tipo_lower:
        asunto = f"Propuesta de página web para {nombre} (para recibir reservas directas)"
        body_text = f"""Hola, equipo de {nombre}, ¿cómo van?

Vi su hospedaje en Google Maps y me pareció un lugar increíble para desconectarse y disfrutar de la naturaleza. Sin embargo, noté que no cuentan con una página web propia donde los viajeros puedan ver las fotos del lugar, consultar servicios y reservar directamente.

Depender solo de redes sociales o pagar comisiones a plataformas externas hace que pierdan reservas directas de turistas que buscan hospedaje por Google.

Me tomé el atrevimiento de crearles un borrador de página web interactivo y adaptado para su alojamiento. Ya está publicada y pueden verla funcionando desde su celular en este enlace:

👉 {demo_url}

📌 Nota sobre este borrador: Esta es una estructura visual preliminar de muestra. Si les gusta la propuesta, la adaptamos y pulimos al 100% con sus fotos en alta resolución, sus tarifas exactas, sus políticas de reserva y todos los ajustes que quieran hacer antes de publicarla oficialmente bajo su dominio.

Mírenla tranquilos sin ningún compromiso. Quedo atento si quieren que la llevemos al siguiente nivel.

¡Un saludo y muchos éxitos con el hospedaje!

Edilberto Sarmiento
Diseñador Web Freelance
📱 WhatsApp: +57 310 481 6153"""

    elif 'fisio' in tipo_lower or 'rehab' in tipo_lower or 'terap' in tipo_lower:
        asunto = f"Una propuesta de página web para {nombre}"
        body_text = f"""Hola, equipo de {nombre}, ¿cómo están?

Estuve analizando servicios de fisioterapia en la ciudad y encontré su perfil con excelentes valoraciones de sus pacientes. Noté que actualmente no cuentan con una página web oficial donde las personas con dolores o lesiones físicas puedan informarse sobre sus terapias y agendar su cita.

Cuando alguien sufre una lesión o dolor agudo, busca rápidamente en Google quién le proporcione confianza inmediata desde el celular.

Para ayudarles a captar más pacientes, les construí un borrador de página web 100% interactivo y personalizado. Pueden revisarlo funcionando en vivo desde su celular aquí:

👉 {demo_url}

📌 Aclaración clave: Esto es un borrador preliminar para mostrarles el potencial visual. Si les atrae la idea, la ajustamos y perfeccionamos al 100% con sus fotos de consultorio, su logo, sus tarifas y todos los cambios que consideren necesarios antes de hacer la publicación oficial en su dominio propio.

Revisen la propuesta sin ningún compromiso. Quedo a su disposición.

Cordialmente,

Edilberto Sarmiento
Diseñador Web Freelance
📱 WhatsApp: +57 310 481 6153"""

    else:
        asunto = f"Diseñé un borrador de página web para {nombre}"
        body_text = f"""Hola, equipo de {nombre}, ¿cómo están?

Estaba revisando empresas locales en Google Maps y encontré su negocio. Vi que tienen muy buen trabajo, pero noté que aún no cuentan con una página web oficial donde sus clientes potenciales puedan conocer todos sus productos o servicios y ponerse en contacto de forma inmediata.

Me tomé el trabajo de prepararles un borrador de página web completamente funcional. Pueden ver cómo luce en vivo desde su celular en el siguiente enlace:

👉 {demo_url}

📌 Es importante resaltar: Este enlace es solo una propuesta preliminar de muestra. Si les resulta de interés, nos sentamos a ajustarla y afinarla al 100% con su información oficial, sus fotos reales, sus textos y todos los requerimientos específicos que necesiten antes de lanzarla en su dominio definitivo.

Pueden revisarla sin ningún compromiso. Quedo a sus órdenes si desean que la llevemos a cabo.

Saludos cordiales,

Edilberto Sarmiento
Diseñador Web Freelance
📱 WhatsApp: +57 310 481 6153"""

    # Formatear HTML básico a partir del texto
    paragraphs = body_text.split('\n\n')
    html_paragraphs_list = []
    for p in paragraphs:
        p_clean = p.replace('\n', '<br>')
        html_paragraphs_list.append(f"<p>{p_clean}</p>")
    html_paragraphs = "".join(html_paragraphs_list)
    body_html = f"""<!DOCTYPE html>
<html>
  <head><meta charset="utf-8"></head>
  <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; color: #1e293b; max-width: 600px; margin: 0 auto; padding: 20px;">
    {html_paragraphs}
  </body>
</html>"""

    return asunto, body_html, body_text

def send_via_mailrelay(recipient_name, recipient_email, subject, body_html, body_text):
    url = f"https://{MAILRELAY_HOST}/api/v1/send_emails"
    
    payload = {
        "from": {
            "email": SENDER_EMAIL,
            "name": SENDER_NAME
        },
        "subject": subject,
        "html_part": body_html,
        "text_part": body_text,
        "to": [
            {
                "email": recipient_email,
                "name": recipient_name
            }
        ]
    }
    
    headers = {
        "Content-Type": "application/json",
        "X-AUTH-TOKEN": MAILRELAY_API_KEY
    }
    
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    
    try:
        with urllib.request.urlopen(req) as response:
            res_body = response.read().decode("utf-8")
            res_json = json.loads(res_body)
            return True, res_json
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8")
        return False, f"HTTP Error {e.code}: {err_body}"
    except Exception as e:
        return False, str(e)

def process_email_sending():
    print("=== PIPELINE 3: ENVÍO REAL DE CORREOS VÍA MAILRELAY API ===")
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: No se encontró {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb['EMPRESAS MAPS']
    rows = list(sheet.iter_rows(values_only=False))

    targets = []
    for idx, r in enumerate(rows[1:], start=2):
        vals = [cell.value for cell in r]
        if not any(vals): continue
        nombre = vals[0]
        tipo = vals[2]
        correo = vals[3]
        estado = str(vals[9]) if len(vals) > 9 and vals[9] else ''

        if 'WEB_BORRADOR_LISTA' in estado or 'PENDIENTE' in estado:
            if correo and str(correo).strip() and '@' in str(correo) and str(correo).strip() != 'N/D':
                demo_url = f"https://betto12680.github.io/borradoes-webs/{nombre}/"
                if 'LINK: ' in estado:
                    demo_url = estado.split('LINK: ')[1].strip()
                targets.append((idx, nombre, tipo, str(correo).strip(), demo_url))

    print(f"Total empresas pendientes para envío de correo: {len(targets)}")
    if not targets:
        print("No hay empresas pendientes para envío de correo.")
        return

    sent_count = 0
    today_str = datetime.now().strftime("%Y-%m-%d")

    for i in range(0, len(targets), BATCH_SIZE):
        batch = targets[i:i + BATCH_SIZE]
        print(f"\n--- Procesando Bloque {i//BATCH_SIZE + 1} ({len(batch)} correos) ---")

        for idx, nombre, tipo, correo, demo_url in batch:
            asunto, body_html, body_text = get_email_template(tipo, nombre, demo_url)
            print(f"✉ Enviando a: {nombre} ({correo})... ", end="", flush=True)
            
            success, resp = send_via_mailrelay(nombre, correo, asunto, body_html, body_text)
            
            if success:
                nuevo_estado = f"CORREO_ENVIADO | FECHA: {today_str} | LINK: {demo_url}"
                sheet.cell(row=idx, column=10, value=nuevo_estado)
                sent_count += 1
                print("✔ ENVIADO VIA MAILRELAY")
            else:
                print(f"❌ ERROR: {resp}")
                
            time.sleep(PAUSE_BETWEEN_EMAILS_SEC)

        wb.save(EXCEL_PATH)

        if i + BATCH_SIZE < len(targets):
            print(f"\n⏸ Pausa anti-spam de {PAUSE_BETWEEN_BATCHES_SEC} segundos antes del siguiente bloque...")
            time.sleep(PAUSE_BETWEEN_BATCHES_SEC)

    print(f"\n✔ PROCESO FINALIZADO: Se enviaron con éxito {sent_count} correos a través de Mailrelay API.")

if __name__ == "__main__":
    process_email_sending()
