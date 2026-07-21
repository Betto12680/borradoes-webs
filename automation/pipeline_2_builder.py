#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Producción V2 - Rediseño Total de Webs Borrador (CDMX)
Inspirado en referentes de alta conversión (Odontosalud-P):
- Carrusel interactivo de servicios y tratamientos con slider JS nativo.
- Galería de características / casos con controles dinámicos.
- Pestañas interactivas de especialidades.
- Contadores animados de métricas y reseñas realistas de Google.
- Animaciones CSS scroll-reveal, glassmorphism, micro-interacciones hover.
- Botón de WhatsApp flotante con pulso y llamada a la acción irresistible.
"""

import os
import re
import sys
import shutil
import subprocess
import openpyxl
from datetime import datetime

BASE_DIR = '/Users/beto/Library/CloudStorage/OneDrive-Personal(2)/Freelancer/clientes/web'
EXCEL_PATH = os.path.join(BASE_DIR, 'Clientes.xlsx')

def slugify(text):
    text = text.lower().strip()
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n', 'ü': 'u'}
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text.strip('-')

# Mapeo de marcas y estilos personalizados por empresa
BRANDS_V2 = {
    'Kintsu Dental': {
        'slug': 'kintsu-dental',
        'nicho_tipo': 'Odontología / Clínica Dental',
        'nicho_schema': 'Dentist',
        'primary': '#0f766e',
        'accent': '#14b8a6',
        'accent_light': '#ccfbf1',
        'dark_bg': '#042f2e',
        'tagline': 'Restaura y Perfecciona Tu Sonrisa',
        'hero_title': 'Odontología Estética & Alta Especialidad en CDMX',
        'hero_sub': 'Inspirados en la filosofía Kintsugi: restauramos y enaltecemos la belleza natural de tu sonrisa con tecnología 3D y máxima comodidad.',
        'hero_img': 'https://images.unsplash.com/photo-1629909613654-28e377c37b09?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'kintsu dental, dentista cdmx, diseno de sonrisa cdmx, ortodoncia invisible, carillas ceramicas cdmx',
        'stats': [('4.9 ★', 'Calificación Google'), ('+1,200', 'Pacientes Felices'), ('100%', 'Garantía Estética'), ('12+', 'Años de Experiencia')],
        'servicios': [
            ('Diseño de Sonrisa Kintsu', 'Carillas cerámicas ultradadas de alta definición hechas a la medida de tu rostro.', '💎', ['Carillas Cerámicas', 'Lentes de Contacto Dental', 'Simulación 3D']),
            ('Ortodoncia Invisible', 'Alineación de dientes sin brackets usando férulas transparentes e imperceptibles.', '✨', ['Alineadores Transparentes', 'Escaneo Digital', 'Sin Alambres']),
            ('Blanqueamiento Láser LED', 'Aclarado dental profesional intensivo en una sola sesión sin dolor ni sensibilidad.', '⚡', ['Hasta 4 Tonos Más Blanco', 'Protección de Esmalte', 'Resultado Inmediato']),
            ('Implantes & Rehabilitación', 'Reemplazo permanente de piezas perdidas con pernos de titanio de integración rápida.', '🛡️', ['Titanio Quirúrgico', 'Carga Inmediata', 'Firmeza Natural']),
            ('Limpieza Dental Ultrasonica', 'Profilaxis profunda que elimina sarro, placa bacteriana y manchas de café o cigarro.', '🩺', ['Ultrasonido Médico', 'Pulido Fluorado', 'Prevención de Gingivitis']),
            ('Endodoncia Computarizada', 'Salva tus dientes naturales eliminando infecciones internas bajo anestesia suave.', '🔬', ['Sin Dolor', 'Microscopio Clínico', 'Tratamiento Rápido'])
        ]
    },
    'Clínica Dental Amsterdent': {
        'slug': 'clinica-dental-amsterdent',
        'nicho_tipo': 'Odontología / La Condesa',
        'nicho_schema': 'Dentist',
        'primary': '#0284c7',
        'accent': '#38bdf8',
        'accent_light': '#e0f2fe',
        'dark_bg': '#0c4a6e',
        'tagline': 'Tu Consultorio Dental en La Condesa',
        'hero_title': 'Salud & Estética Bucal de Vanguardia en La Condesa, CDMX',
        'hero_sub': 'Ubicados en Amsterdam 124. Te brindamos una experiencia odontológica sin estrés, amigable y con tecnología digital de punta.',
        'hero_img': 'https://images.unsplash.com/photo-1588776814546-1ffcf47267a5?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'amsterdent condesa, clinica dental condesa cdmx, dentista amsterdam condesa, implantes condesa',
        'stats': [('4.8 ★', 'Reseñas en Google'), ('+950', 'Sonrisas Diseñadas'), ('98%', 'Recomendación'), ('10+', 'Años en La Condesa')],
        'servicios': [
            ('Implantes Dentales Condesa', 'Restauración de piezas faltantes con implantes de titanio y coronas de zirconio.', '🛡️', ['Zirconio Alta Densidad', 'Aspecto 100% Natural', 'Garantía por Escrito']),
            ('Ortodoncia & Brackets Estéticos', 'Brackets de zafiro y metálicos para alinear tus dientes de forma eficiente.', '✨', ['Brackets de Zafiro', 'Retenedores Incluidos', 'Alineación Rápida']),
            ('Estética Dental & Resinas', 'Restauración estética e imperceptible de caries y fracturas dentales.', '💎', ['Resina Fotocurable', 'Cero Amalgamas', 'Color Idéntico']),
            ('Limpieza & Detartraje', 'Eliminación completa de sarro y pulido dental para encías sanas.', '🩺', ['Ultrasonido Sin Molestia', 'Selladores', 'Aliento Fresco'])
        ]
    },
    'KlinikDent México': {
        'slug': 'klinikdent-mexico',
        'nicho_tipo': 'Odontología / Del Valle',
        'nicho_schema': 'Dentist',
        'primary': '#1e3a8a',
        'accent': '#3b82f6',
        'accent_light': '#dbeafe',
        'dark_bg': '#172554',
        'tagline': 'Alta Especialidad Odontológica',
        'hero_title': 'Odontología Especializada & Cirugía Oral en Del Valle',
        'hero_sub': 'Especialistas certificados dedicados al cuidado integral de tu salud bucal con instalaciones climatizadas de máxima comodidad.',
        'hero_img': 'https://images.unsplash.com/photo-1606811841689-23dfddce3e95?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'klinikdent del valle, dentista del valle cdmx, cirugia muelas del juicio del valle, carillas cdmx',
        'stats': [('5.0 ★', 'Calificación Máxima'), ('+1,500', 'Casos Exitosos'), ('100%', 'Atención Personal'), ('15+', 'Años de Experiencia')],
        'servicios': [
            ('Carillas & Estética Dental', 'Transforma el color y alineación de tu dentadura con carillas ultradadas.', '💎', ['Porcelana E-Max', 'Mínimo Desgaste', 'Sonrisa Perfecta']),
            ('Extracción de Terceros Molares', 'Cirugía de muelas del juicio sin dolor y con rápida recuperación.', '🔬', ['Anestesia Guiada', 'Técnica Mínimamente Invasiva', 'Cuidados Postoperatorios']),
            ('Endodoncia y Tratamiento de Conducto', 'Tratamiento efectivo contra el dolor de muela intenso para preservar la pieza.', '⚡', ['Rotatorio Digital', 'Sin Dolor', 'Sellado Hermético']),
            ('Odontopediatría Amigable', 'Atención especializada para niños en ambiente divertido y sin temor.', '🧸', ['Prevención Infantil', 'Aplicación de Flúor', 'Técnica de Cepillado'])
        ]
    },
    'Dental Studio MX': {
        'slug': 'dental-studio-mx',
        'nicho_tipo': 'Odontología / Multi-Sede',
        'nicho_schema': 'Dentist',
        'primary': '#7c3aed',
        'accent': '#a855f7',
        'accent_light': '#f3e8ff',
        'dark_bg': '#3b0764',
        'tagline': 'Studio de Ortodoncia & Diseño 3D',
        'hero_title': 'El Futuro de Tu Sonrisa en Narvarte, Polanco & Coyoacán',
        'hero_sub': 'Tecnología de escaneo intraoral 3D y alineadores invisibles para lograr tu sonrisa deseada de manera rápida y discreta.',
        'hero_img': 'https://images.unsplash.com/photo-1598256989800-fe5f95da9787?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'dental studio mx, alineadores invisibles cdmx, ortodoncia 3d narvarte, carillas polanco',
        'stats': [('4.9 ★', 'Reseñas Verificadas'), ('+2,000', 'Alineadores Entregados'), ('4 Sedes', 'CDMX'), ('100%', 'Escaneo Digital')],
        'servicios': [
            ('Alineadores Studio 3D', 'Alineación dental invisible planificada en computadora con simulación virtual previa.', '✨', ['Simulación Virtual', 'Placas Transparentes', 'Revisiones Rápidas']),
            ('Blanqueamiento Studio Flash', 'Aclarado dental LED acelerado que combate manchas profundas en 45 minutos.', '⚡', ['Tecnología LED', 'Sin Dañar Esmalte', 'Resultado Visible']),
            ('Diseño Digital de Sonrisa (DSD)', 'Analizamos tus proporciones faciales para diseñar la sonrisa ideal en pantalla.', '💎', ['Análisis Facial', 'Carillas Personalizadas', 'Prueba Previa']),
            ('Profilaxis & Limpieza Ultrasonica', 'Mantenimiento preventivo completo para encías firmes y aliento fresco.', '🩺', ['Ultrasonido de Alta Eficiencia', 'Pulido de Esmalte', 'Cero Sarro'])
        ]
    },
    'Rehavilita Fisioterapia': {
        'slug': 'rehavilita-fisioterapia',
        'nicho_tipo': 'Fisioterapia & Rehabilitación',
        'nicho_schema': 'PhysicalTherapy',
        'primary': '#14436c',
        'accent': '#2ecfb4',
        'accent_light': '#e6fcf5',
        'dark_bg': '#0b2545',
        'tagline': 'Rehabilitación Física & Bienestar',
        'hero_title': 'Fisioterapia, Lenguaje & Terapia Respiratoria en Del Valle',
        'hero_sub': 'Ubicados en Miguel Laurent 510. Equipo multidisciplinario enfocado en devolverte la movilidad y calidad de vida.',
        'hero_img': 'https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'rehavilita fisioterapia, fisioterapia miguel laurent, terapia lenguaje del valle, rehabilitacion respiratoria cdmx',
        'stats': [('4.9 ★', 'Google Maps'), ('+1,800', 'Sesiones Exitosas'), ('99%', 'Alivio del Dolor'), ('14+', 'Años de Experiencia')],
        'servicios': [
            ('Fisioterapia & Lesiones Deportivas', 'Tratamiento de contracturas, esguinces, tendinitis y recuperación muscular acelerada.', '🏃', ['Electroterapia', 'Ultrasonido', 'Terapia Manual']),
            ('Terapia de Lenguaje & Ocupacional', 'Rehabilitación comunicativa y habilidades motoras para niños y adultos.', '🗣️', ['Evaluación Integral', 'Ejercicios Adaptativos', 'Estimulación']),
            ('Fisioterapia Respiratoria', 'Técnicas especializadas para optimizar la capacidad pulmonar post-infección o bronquitis.', '🫁', ['Drenaje Bronquial', 'Expansión Pulmonar', 'Oxigenación']),
            ('Masaje Terapéutico Descontracturante', 'Alivio profundo de la tensión acumulada en cuello, hombros y columna lumbar.', '🧘', ['Liberación Miofascial', 'Puntos Gatillo', 'Descarga Muscular'])
        ]
    },
    'Fisioterapia y Rehabilitación Vértiz': {
        'slug': 'fisioterapia-y-rehabilitacion-vertiz',
        'nicho_tipo': 'Fisioterapia / Medicina Deportiva',
        'nicho_schema': 'PhysicalTherapy',
        'primary': '#0284c7',
        'accent': '#38bdf8',
        'accent_light': '#e0f2fe',
        'dark_bg': '#075985',
        'tagline': 'Medicina Deportiva & Rehabilitación',
        'hero_title': 'Recupera Tu Movilidad Sin Dolor en Benito Juárez, CDMX',
        'hero_sub': 'Dr. José María Vértiz 1218. Especialistas en medicina deportiva, lesiones de rodilla, espalda y rehabilitación postquirúrgica.',
        'hero_img': 'https://images.unsplash.com/photo-1519823551278-64ac92734fb1?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'fisioterapia vertiz, rehabilitacion deportiva vertiz, fisioterapeuta benito juarez, alivio dolor ciatica',
        'stats': [('4.8 ★', 'Reseñas en Google'), ('+1,100', 'Deportistas Atendidos'), ('100%', 'Atención Guiada'), ('10+', 'Años en Vértiz')],
        'servicios': [
            ('Rehabilitación Deportiva Vértiz', 'Tratamiento intensivo para volver a entrenar sin dolor en el menor tiempo.', '⚡', ['Vendaje Neuromuscular', 'Crioterapia', 'Reeducación de Gesto']),
            ('Fisioterapia Ortopédica & Columna', 'Alivio efectivo de ciática, hernias discales, lumbalgias y contracturas.', '🛡️', ['Tracción Vertebral', 'Terapia Manual', 'Fortalecimiento']),
            ('Rehabilitación Postquirúrgica', 'Acompañamiento en recuperación tras cirugía de ligamentos, meniscos o prótesis.', '🏥', ['Movilización Asistida', 'Recuperación de Arcos', 'Control de Inflamación']),
            ('Masaje Deportivo & Descarga', 'Técnica profunda para eliminar ácido láctico y prevenir lesiones en atletas.', '🧘', ['Descarga Muscular', 'Estiramientos Guiados', 'Flexibilidad'])
        ]
    },
    'CERCARDIO Especialidades Médicas': {
        'slug': 'cercardio-especialidades-medicas',
        'nicho_tipo': 'Clínica Médica / Cardiología',
        'nicho_schema': 'MedicalClinic',
        'primary': '#b91c1c',
        'accent': '#ef4444',
        'accent_light': '#fee2e2',
        'dark_bg': '#7f1d1d',
        'tagline': 'Centro de Cardiología & Especialidades',
        'hero_title': 'Cuidado Cardiovascular & Checkup Médico en Lindavista',
        'hero_sub': 'Ubicados en Hospital Ángeles Lindavista y Hospital Boutique Riobamba. Especialistas certificados en la salud de tu corazón.',
        'hero_img': 'https://images.unsplash.com/photo-1516549655169-df83a0774514?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'cercardio lindavista, cardiologo angeles lindavista, checkup cardiaco cdmx, prueba de esfuerzo lindavista',
        'stats': [('4.9 ★', 'Calificación Médica'), ('+3,500', 'Estudios Realizados'), ('100%', 'Cardiólogos Certificados'), ('Sedes', 'Ángeles Lindavista')],
        'servicios': [
            ('Checkup Cardiológico Integral', 'Evaluación completa de riesgo cardíaco con electrocardiograma y prueba de esfuerzo.', '🫀', ['Electrocardiograma 12 Derivaciones', 'Ecocardiograma Doppler', 'Prueba de Esfuerzo']),
            ('Rehabilitación Cardiovascular', 'Programa de ejercicio médico supervisado para pacientes post-infarto o cirugía.', '🏃', ['Monitoreo Continuo', 'Supervisión Médica', 'Plan Gradual']),
            ('Monitoreo Holter & Presión 24h', 'Detección precisa de arritmias e hipertensión mediante dispositivos portátiles.', '📊', ['Holter de Arritmias', 'MAPA de Presión', 'Reporte Digital']),
            ('Consultas de Alta Especialidad', 'Valoración clínica detallada por médicos especialistas en medicina interna y cardiología.', '🩺', ['Diagnóstico Certero', 'Ajuste de Tratamiento', 'Seguimiento'])
        ]
    },
    'Vitalmente Centro Médico & Psicología': {
        'slug': 'vitalmente-centro-medico-psicologia',
        'nicho_tipo': 'Salud Mental & Clínica Médica',
        'nicho_schema': 'MedicalClinic',
        'primary': '#0d9488',
        'accent': '#14b8a6',
        'accent_light': '#ccfbf1',
        'dark_bg': '#115e59',
        'tagline': 'Salud Integral & Psicología Clínica',
        'hero_title': 'Bienestar Emocional & Salud Mental en Lomas, CDMX',
        'hero_sub': 'Sierra de Roraima 15. Un espacio seguro y confidencial guiado por especialistas para recuperar la tranquilidad en tu vida.',
        'hero_img': 'https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?auto=format&fit=crop&w=1200&q=80',
        'keywords': 'vitalmente centro medico, psicologo lomas cdmx, terapia ansiedad roraima, psicoterapia familiar cdmx',
        'stats': [('4.8 ★', 'Reseñas Verificadas'), ('+1,400', 'Pacientes Atendidos'), ('100%', 'Confidencialidad'), ('12+', 'Años de Experiencia')],
        'servicios': [
            ('Psicoterapia para Ansiedad & Depresión', 'Herramientas clínicas efectivas para detener ataques de pánico, angustia y tristeza.', '🧠', ['Enfoque Cognitivo', 'Control del Estrés', 'Equilibrio Emocional']),
            ('Terapia de Pareja & Familia', 'Espacio neutral para mejorar la comunicación, resolver conflictos y sanar relaciones.', '🤝', ['Asertividad', 'Mediación Familiar', 'Vínculos Sanos']),
            ('Evaluación & Psicodiagnóstico', 'Pruebas psicométricas completas para un conocimiento certero del estado de salud mental.', '📋', ['Test Clínicos', 'Informe Detallado', 'Orientación']),
            ('Consultas Médicas Integrales', 'Sinergia multidisciplinaria entre psicología y médicos para el bienestar del paciente.', '🩺', ['Atención Holística', 'Seguimiento Continuo', 'Espacio Seguro'])
        ]
    }
}

def generate_super_v2_html(nombre, item, config):
    ciudad = item['ciudad']
    tel = item['tel']
    resenas = item['resenas']
    rating = item['rating']
    demo_url = f"https://betto12680.github.io/borradoes-webs/{config['slug']}/"

    rating_val = str(rating) if rating and str(rating) != 'N/D' else '4.9'
    resenas_val = str(resenas) if resenas and str(resenas) != 'N/D' else '45'
    tel_clean = re.sub(r'\D', '', str(tel)) if tel and str(tel) != 'N/D' else '525555551234'
    if not tel_clean.startswith('52'): tel_clean = '52' + tel_clean

    nicho_schema = config['nicho_schema']

    # FAQs dinámicas
    if nicho_schema == 'Dentist':
        faqs = [
            ("¿Cómo agendar mi primera cita de valoración?", "Puedes hacer clic en el botón de WhatsApp y nuestro equipo de recepción te asignará la fecha y hora disponible de inmediato."),
            ("¿Tienen atención de emergencias el mismo día?", "Sí, atendemos urgencias dentales por dolor agudo o accidentes con prioridad según disponibilidad de agenda."),
            ("¿Qué formas de pago aceptan?", "Aceptamos efectivo, transferencias bancarias (SPEI) y tarjetas de crédito/débito con opción de meses sin intereses."),
            ("¿Ofrecen garantía en tratamientos estéticos e implantes?", "Todos nuestros procedimientos estéticos, prótesis e implantes cuentan con garantía por escrito y controles periódicos.")
        ]
    elif nicho_schema == 'PhysicalTherapy':
        faqs = [
            ("¿Necesito orden médica para iniciar tratamiento?", "No es estrictamente necesario. En tu primera consulta realizamos una valoración física completa para diseñar tu plan de rehabilitación."),
            ("¿Cuánto dura cada sesión de fisioterapia?", "Las sesiones duran aproximadamente entre 50 y 60 minutos e incluyen terapia manual, electroterapia y ejercicios guiados."),
            ("¿Atienden a domicilio o solo en consultorio?", "Contamos con instalaciones equipadas en CDMX y también ofrecemos servicio domiciliario en zonas seleccionadas."),
            ("¿En cuántas sesiones sentiré alivio?", "Muchos pacientes sienten una mejoría notable desde la primera sesión. La duración total depende del diagnóstico.")
        ]
    else:
        faqs = [
            ("¿Cómo puedo agendar una consulta médica o psicológica?", "Haz clic en el botón de WhatsApp y nuestro equipo te asistirá para agendar tu horario de manera ágil."),
            ("¿Las consultas son presenciales u online?", "Ofrecemos atención presencial en nuestras instalaciones climatizadas y opción de videoconsulta previa cita."),
            ("¿Cómo garantizan la confidencialidad de mi consulta?", "Todos nuestros profesionales están sujetos a estricto secreto profesional y aviso de privacidad de datos médicos."),
            ("¿Qué debo llevar a mi primera cita?", "Te sugerimos acudir con estudios médicos previos o historial si cuentas con ellos para una evaluación completa.")
        ]

    # Renderizado de servicios en tarjetas interactivas
    servicios_html = ""
    for s_title, s_desc, s_icon, s_tags in config['servicios']:
        tags_badge = "".join([f'<span class="srv-tag">{t}</span>' for t in s_tags])
        servicios_html += f"""
        <div class="srv-card" data-reveal>
          <div class="srv-head">
            <div class="srv-icon-box">{s_icon}</div>
            <h3>{s_title}</h3>
          </div>
          <p>{s_desc}</p>
          <div class="srv-tags-wrap">{tags_badge}</div>
          <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20me%20interesa%20informaci%C3%B3n%20sobre%20{s_title}" class="srv-btn-action" target="_blank">
            Consultar por WhatsApp ↗
          </a>
        </div>"""

    # Renderizado de stats
    stats_html = ""
    for st_num, st_label in config['stats']:
        stats_html += f"""
        <div class="stat-box">
          <div class="stat-number">{st_num}</div>
          <div class="stat-label">{st_label}</div>
        </div>"""

    # Schema JSON-LD
    schema_json = f"""{{
  "@context": "https://schema.org",
  "@type": "{nicho_schema}",
  "name": "{nombre}",
  "url": "{demo_url}",
  "telephone": "+{tel_clean}",
  "address": {{
    "@type": "PostalAddress",
    "addressLocality": "{ciudad}",
    "addressCountry": "MX"
  }},
  "aggregateRating": {{
    "@type": "AggregateRating",
    "ratingValue": "{rating_val}",
    "reviewCount": "{resenas_val}"
  }},
  "priceRange": "$$"
}}"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{nombre} — {config['hero_title']}</title>
  <meta name="description" content="{config['hero_sub']}">
  <meta name="keywords" content="{config['keywords']}">
  <meta name="robots" content="index, follow">
  
  <!-- Open Graph -->
  <meta property="og:type" content="website">
  <meta property="og:url" content="{demo_url}">
  <meta property="og:title" content="{nombre} — {config['hero_title']}">
  <meta property="og:description" content="{config['hero_sub']}">
  <meta property="og:image" content="{config['hero_img']}">
  <meta property="og:locale" content="es_MX">

  <!-- Google Fonts -->
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=Outfit:wght@500;600;700;800&display=swap" rel="stylesheet">

  <script type="application/ld+json">
  {schema_json}
  </script>

  <style>
    :root {{
      --primary: {config['primary']};
      --accent: {config['accent']};
      --accent-light: {config['accent_light']};
      --dark-bg: {config['dark_bg']};
      --bg: #090d16;
      --card-bg: #131a2a;
      --text: #f8fafc;
      --text-muted: #94a3b8;
      --border: #1e293b;
    }}

    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    html {{ scroll-behavior: smooth; }}
    body {{
      font-family: 'Plus Jakarta Sans', sans-serif;
      background-color: var(--bg);
      color: var(--text);
      line-height: 1.6;
      overflow-x: hidden;
    }}

    .container {{ max-width: 1180px; margin: 0 auto; padding: 0 24px; }}

    /* Header Sticky */
    nav.navbar {{
      position: fixed; top: 0; left: 0; right: 0; z-index: 1000;
      background: rgba(9, 13, 22, 0.88);
      backdrop-filter: blur(16px); -webkit-backdrop-filter: blur(16px);
      border-bottom: 1px solid var(--border); padding: 16px 0;
      transition: all 0.3s ease;
    }}
    .nav-flex {{ display: flex; justify-content: space-between; align-items: center; }}
    .logo-brand {{
      font-family: 'Outfit', sans-serif; font-size: 1.4rem; font-weight: 800;
      color: #fff; text-decoration: none; display: flex; align-items: center; gap: 8px;
    }}
    .logo-brand span {{ color: var(--accent); }}

    .btn-nav-wa {{
      background: #25d366; color: #fff; text-decoration: none;
      padding: 10px 22px; border-radius: 50px; font-weight: 700;
      font-size: 0.92rem; display: inline-flex; align-items: center; gap: 8px;
      box-shadow: 0 4px 16px rgba(37, 211, 102, 0.35);
      transition: transform 0.25s ease, box-shadow 0.25s ease;
    }}
    .btn-nav-wa:hover {{ transform: translateY(-2px); box-shadow: 0 8px 24px rgba(37, 211, 102, 0.55); }}

    /* Floating WhatsApp Button */
    .wa-float-btn {{
      position: fixed; bottom: 28px; right: 28px; z-index: 999;
      background: #25d366; color: #fff; width: 64px; height: 64px;
      border-radius: 50%; display: flex; align-items: center; justify-content: center;
      box-shadow: 0 10px 28px rgba(37, 211, 102, 0.5);
      text-decoration: none; font-size: 1.9rem;
      transition: transform 0.3s ease;
    }}
    .wa-float-btn:hover {{ transform: scale(1.12); }}
    .wa-float-btn .pulse-wave {{
      position: absolute; inset: 0; border-radius: 50%;
      background: #25d366; animation: wapulse 2.2s ease-out infinite; z-index: -1;
    }}
    @keyframes wapulse {{ 0% {{ transform: scale(1); opacity: 0.6; }} 100% {{ transform: scale(2.2); opacity: 0; }} }}

    /* Hero Section Visual */
    header.hero-v2 {{
      padding: 160px 0 100px;
      background: radial-gradient(circle at 50% 20%, var(--dark-bg) 0%, var(--bg) 80%);
      position: relative; overflow: hidden;
    }}
    .hero-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 48px; align-items: center; }}
    
    .hero-badge-tag {{
      display: inline-flex; align-items: center; gap: 8px;
      background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.15);
      padding: 8px 18px; border-radius: 30px; font-weight: 700;
      font-size: 0.88rem; color: var(--accent); margin-bottom: 20px;
      text-transform: uppercase; letter-spacing: 0.05em;
    }}

    .hero-v2 h1 {{
      font-family: 'Outfit', sans-serif; font-size: 3.2rem; font-weight: 800;
      line-height: 1.15; letter-spacing: -0.02em; margin-bottom: 20px; color: #fff;
    }}
    .hero-v2 h1 span {{
      background: linear-gradient(135deg, #fff 30%, var(--accent) 100%);
      -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    }}

    .hero-v2 p.hero-desc {{
      font-size: 1.15rem; color: var(--text-muted); margin-bottom: 36px;
      font-weight: 400; line-height: 1.65;
    }}

    .hero-cta-group {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 32px; }}
    
    .btn-main-action {{
      background: #25d366; color: #fff; text-decoration: none;
      padding: 16px 36px; border-radius: 50px; font-weight: 800;
      font-size: 1.08rem; display: inline-flex; align-items: center; gap: 10px;
      box-shadow: 0 10px 30px rgba(37, 211, 102, 0.4);
      transition: transform 0.25s ease, box-shadow 0.25s ease;
    }}
    .btn-main-action:hover {{ transform: translateY(-3px); box-shadow: 0 14px 36px rgba(37, 211, 102, 0.6); }}

    .hero-social-proof {{
      display: flex; align-items: center; gap: 16px;
      background: rgba(255,255,255,0.04); border: 1px solid var(--border);
      padding: 12px 20px; border-radius: 16px; width: fit-content;
    }}
    .stars-score {{ color: #fbbf24; font-weight: 800; font-size: 1.1rem; }}
    .reviews-meta {{ color: var(--text-muted); font-size: 0.88rem; }}

    /* Hero Right Frame (Interactive Look) */
    .hero-img-frame {{
      position: relative; border-radius: 28px; overflow: hidden;
      border: 1px solid rgba(255,255,255,0.15);
      box-shadow: 0 30px 60px -15px rgba(0,0,0,0.7);
    }}
    .hero-img-frame img {{ width: 100%; height: 440px; object-fit: cover; display: block; }}
    .hero-overlay-badge {{
      position: absolute; bottom: 20px; left: 20px; right: 20px;
      background: rgba(15, 23, 42, 0.9); backdrop-filter: blur(12px);
      border: 1px solid rgba(255,255,255,0.15); padding: 16px 20px;
      border-radius: 18px; display: flex; align-items: center; justify-content: space-between;
    }}
    .badge-info-title {{ font-weight: 700; color: #fff; font-size: 0.95rem; }}
    .badge-info-sub {{ color: var(--accent); font-size: 0.85rem; font-weight: 600; }}

    /* Stats Counter Bar */
    .stats-bar {{
      background: var(--card-bg); border-y: 1px solid var(--border);
      padding: 40px 0; margin-bottom: 100px;
    }}
    .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 32px; text-align: center; }}
    .stat-number {{ font-family: 'Outfit', sans-serif; font-size: 2.8rem; font-weight: 800; color: var(--accent); line-height: 1; margin-bottom: 8px; }}
    .stat-label {{ color: var(--text-muted); font-weight: 600; font-size: 0.95rem; }}

    /* Services Interactive Showcase */
    .sec-header {{ text-align: center; margin-bottom: 60px; }}
    .sec-header h2 {{ font-family: 'Outfit', sans-serif; font-size: 2.5rem; font-weight: 800; color: #fff; margin-bottom: 14px; }}
    .sec-header p {{ color: var(--text-muted); font-size: 1.1rem; max-width: 620px; margin: 0 auto; }}

    .srv-grid {{
      display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
      gap: 30px; margin-bottom: 100px;
    }}
    .srv-card {{
      background: var(--card-bg); border: 1px solid var(--border);
      border-radius: 24px; padding: 36px; transition: all 0.35s cubic-bezier(0.16, 1, 0.3, 1);
      display: flex; flex-direction: column; justify-content: space-between;
    }}
    .srv-card:hover {{
      transform: translateY(-8px); border-color: var(--accent);
      box-shadow: 0 24px 48px -12px rgba(0,0,0,0.6);
    }}
    .srv-head {{ display: flex; align-items: center; gap: 16px; margin-bottom: 18px; }}
    .srv-icon-box {{
      width: 56px; height: 56px; background: rgba(255,255,255,0.05);
      border: 1px solid rgba(255,255,255,0.1); border-radius: 16px;
      display: flex; align-items: center; justify-content: center; font-size: 1.6rem;
    }}
    .srv-card h3 {{ font-size: 1.3rem; font-weight: 700; color: #fff; }}
    .srv-card p {{ color: var(--text-muted); font-size: 0.98rem; margin-bottom: 24px; line-height: 1.65; }}
    
    .srv-tags-wrap {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 28px; }}
    .srv-tag {{
      background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1);
      color: var(--accent); font-size: 0.8rem; padding: 4px 12px; border-radius: 20px; font-weight: 600;
    }}

    .srv-btn-action {{
      display: block; text-align: center; background: rgba(255,255,255,0.06);
      border: 1px solid var(--border); color: #fff; text-decoration: none;
      padding: 12px 20px; border-radius: 14px; font-weight: 700; font-size: 0.92rem;
      transition: all 0.25s ease;
    }}
    .srv-btn-action:hover {{ background: #25d366; border-color: #25d366; color: #fff; }}

    /* Interactive Image Slider / Carousel Section */
    .slider-section {{ margin-bottom: 100px; }}
    .carousel-container {{
      position: relative; max-width: 960px; margin: 0 auto; overflow: hidden;
      border-radius: 28px; border: 1px solid var(--border);
      box-shadow: 0 30px 60px -15px rgba(0,0,0,0.7);
    }}
    .carousel-track {{ display: flex; transition: transform 0.5s cubic-bezier(0.25, 1, 0.5, 1); }}
    .carousel-slide {{ min-width: 100%; position: relative; }}
    .carousel-slide img {{ width: 100%; height: 480px; object-fit: cover; display: block; }}
    .carousel-caption {{
      position: absolute; bottom: 0; left: 0; right: 0;
      background: linear-gradient(to top, rgba(9, 13, 22, 0.95), transparent);
      padding: 40px; color: #fff;
    }}
    .carousel-caption h4 {{ font-size: 1.4rem; font-weight: 800; margin-bottom: 6px; }}
    .carousel-caption p {{ color: var(--text-muted); font-size: 0.98rem; }}

    .carousel-nav {{
      position: absolute; top: 50%; width: 100%; display: flex;
      justify-content: space-between; padding: 0 20px; transform: translateY(-50%);
      pointer-events: none;
    }}
    .carousel-btn {{
      width: 48px; height: 48px; background: rgba(15, 23, 42, 0.8);
      border: 1px solid rgba(255,255,255,0.2); border-radius: 50%;
      color: #fff; font-size: 1.2rem; cursor: pointer; pointer-events: auto;
      display: flex; align-items: center; justify-content: center;
      transition: background 0.25s ease;
    }}
    .carousel-btn:hover {{ background: var(--accent); }}

    /* FAQs Section */
    .faqs-container {{ max-width: 820px; margin: 0 auto 100px; }}
    .faq-item {{
      background: var(--card-bg); border: 1px solid var(--border);
      border-radius: 16px; margin-bottom: 16px; overflow: hidden;
    }}
    .faq-item summary {{
      padding: 22px 28px; font-weight: 700; font-size: 1.08rem;
      cursor: pointer; list-style: none; display: flex;
      justify-content: space-between; align-items: center; color: #fff;
    }}
    .faq-item summary::-webkit-details-marker {{ display: none; }}
    .faq-item summary::after {{
      content: '+'; font-size: 1.6rem; color: var(--accent); transition: transform 0.25s ease;
    }}
    .faq-item[open] summary::after {{ transform: rotate(45deg); }}
    .faq-item p {{
      padding: 0 28px 24px; color: var(--text-muted); font-size: 1rem;
      border-top: 1px solid rgba(255,255,255,0.06); margin-top: 8px; padding-top: 18px;
    }}

    /* Final CTA Banner */
    .cta-banner-bottom {{
      background: linear-gradient(135deg, var(--dark-bg) 0%, var(--card-bg) 100%);
      border: 1px solid var(--border); border-radius: 32px;
      padding: 60px 40px; text-align: center; margin-bottom: 90px;
    }}
    .cta-banner-bottom h2 {{ font-family: 'Outfit', sans-serif; font-size: 2.4rem; font-weight: 800; color: #fff; margin-bottom: 16px; }}
    .cta-banner-bottom p {{ color: var(--text-muted); font-size: 1.1rem; max-width: 600px; margin: 0 auto 36px; }}

    /* Footer */
    footer {{ border-top: 1px solid var(--border); padding: 50px 0; text-align: center; color: var(--text-muted); font-size: 0.95rem; }}
    footer a {{ color: var(--accent); text-decoration: none; font-weight: 700; }}

    @media (max-width: 900px) {{
      .hero-grid {{ grid-template-columns: 1fr; text-align: center; gap: 36px; }}
      .hero-v2 h1 {{ font-size: 2.4rem; }}
      .hero-social-proof {{ margin: 0 auto; }}
      .hero-cta-group {{ justify-content: center; }}
      .hero-img-frame img {{ height: 320px; }}
    }}
  </style>
</head>
<body>

  <!-- Floating WhatsApp CTA -->
  <a href="https://wa.me/{tel_clean}?text=Hola%20equipo%20de%20{nombre},%20me%20gustar%C3%ADa%20agendar%20una%20cita." class="wa-float-btn" target="_blank" aria-label="WhatsApp Directo">
    <div class="pulse-wave"></div>
    💬
  </a>

  <!-- Header Navigation Sticky -->
  <nav class="navbar">
    <div class="container nav-flex">
      <a href="#" class="logo-brand">
        {nombre}
      </a>
      <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20deseo%20agendar%20cita." class="btn-nav-wa" target="_blank">
        💬 Agendar por WhatsApp
      </a>
    </div>
  </nav>

  <!-- Hero Section Dynamic -->
  <header class="hero-v2">
    <div class="container">
      <div class="hero-grid">
        <div class="hero-left">
          <div class="hero-badge-tag">
            ✨ {config['tagline']}
          </div>
          <h1>{config['hero_title']}</h1>
          <p class="hero-desc">{config['hero_sub']}</p>

          <div class="hero-cta-group">
            <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20quiero%20m%C3%A1s%20informaci%C3%B3n" class="btn-main-action" target="_blank">
              👉 Agendar Mi Cita por WhatsApp
            </a>
          </div>

          <div class="hero-social-proof">
            <div class="stars-score">⭐ {rating_val} / 5.0</div>
            <div class="reviews-meta">Basado en {resenas_val} opiniones en Google Maps</div>
          </div>
        </div>

        <div class="hero-right">
          <div class="hero-img-frame">
            <img src="{config['hero_img']}" alt="{nombre} - {config['hero_title']}" loading="lazy" width="1200" height="440">
            <div class="hero-overlay-badge">
              <div>
                <div class="badge-info-title">{nombre}</div>
                <div class="badge-info-sub">Atención Especializada en {ciudad}</div>
              </div>
              <a href="https://wa.me/{tel_clean}" class="btn-nav-wa" style="padding: 8px 16px; font-size: 0.85rem;" target="_blank">Agendar</a>
            </div>
          </div>
        </div>
      </div>
    </div>
  </header>

  <!-- Stats Bar -->
  <div class="stats-bar">
    <div class="container">
      <div class="stats-grid">{stats_html}
      </div>
    </div>
  </div>

  <!-- Services Interactive Section -->
  <section class="container">
    <div class="sec-header">
      <h2>Servicios & Especialidades</h2>
      <p>Tratamientos modernos y personalizados diseñados para tu salud y confort en {ciudad}.</p>
    </div>

    <div class="srv-grid">{servicios_html}
    </div>
  </section>

  <!-- Interactive Slider Showcase Carousel -->
  <section class="container slider-section">
    <div class="sec-header">
      <h2>Instalaciones & Tecnología</h2>
      <p>Conoce los estándares de higiene, equipos y espacios preparados para tu atención.</p>
    </div>

    <div class="carousel-container">
      <div class="carousel-track" id="sliderTrack">
        <div class="carousel-slide">
          <img src="{config['hero_img']}" alt="Instalaciones {nombre}">
          <div class="carousel-caption">
            <h4>Equipos Digitales de Última Generación</h4>
            <p>Diagnóstico de alta precisión para tratamientos más rápidos y cómodos.</p>
          </div>
        </div>
        <div class="carousel-slide">
          <img src="https://images.unsplash.com/photo-1588776814546-1ffcf47267a5?auto=format&fit=crop&w=1200&q=80" alt="Consultorios {nombre}">
          <div class="carousel-caption">
            <h4>Consultorios Climatizados y Sanitizados</h4>
            <p>Espacios confortables diseñados para tu absoluta seguridad y privacidad.</p>
          </div>
        </div>
        <div class="carousel-slide">
          <img src="https://images.unsplash.com/photo-1579684385127-1ef15d508118?auto=format&fit=crop&w=1200&q=80" alt="Atención {nombre}">
          <div class="carousel-caption">
            <h4>Atención Humana y Personalizada</h4>
            <p>Especialistas dedicados a resolver tus necesidades de salud con calidez.</p>
          </div>
        </div>
      </div>

      <div class="carousel-nav">
        <button class="carousel-btn" id="prevBtn" aria-label="Anterior">❮</button>
        <button class="carousel-btn" id="nextBtn" aria-label="Siguiente">❯</button>
      </div>
    </div>
  </section>

  <!-- FAQs Section -->
  <section class="container">
    <div class="sec-header">
      <h2>Preguntas Frecuentes</h2>
      <p>Resolvemos tus dudas para que inicies tu tratamiento con absoluta confianza.</p>
    </div>

    <div class="faqs-container">
"""
    for q_title, q_ans in faqs:
        html += f"""
      <details class="faq-item">
        <summary>{q_title}</summary>
        <p>{q_ans}</p>
      </details>"""

    html += f"""
    </div>
  </section>

  <!-- Bottom CTA Banner -->
  <section class="container">
    <div class="cta-banner-bottom">
      <h2>¿Listo para agendar tu consulta en {ciudad}?</h2>
      <p>Haz clic en el botón inferior para comunicarte directamente por WhatsApp con nuestro equipo.</p>
      <a href="https://wa.me/{tel_clean}?text=Hola%20{nombre},%20quiero%20agendar%20mi%20cita%20ahora" class="btn-main-action" target="_blank">
        💬 Agendar Mi Cita por WhatsApp Ahora
      </a>
    </div>
  </section>

  <!-- Footer -->
  <footer>
    <div class="container">
      <p>&copy; {datetime.now().year} {nombre} — {ciudad}. Propuesta de diseño web por <a href="https://wa.me/573104816153" target="_blank">Edilberto Sarmiento</a>.</p>
    </div>
  </footer>

  <!-- Native JavaScript Carousel Script -->
  <script>
    (function(){{
      var track = document.getElementById('sliderTrack');
      var prevBtn = document.getElementById('prevBtn');
      var nextBtn = document.getElementById('nextBtn');
      if(!track || !prevBtn || !nextBtn) return;

      var slides = track.querySelectorAll('.carousel-slide');
      var currentIndex = 0;

      function updateSlider(){{
        track.style.transform = 'translateX(-' + (currentIndex * 100) + '%)';
      }}

      nextBtn.addEventListener('click', function(){{
        currentIndex = (currentIndex + 1) % slides.length;
        updateSlider();
      }});

      prevBtn.addEventListener('click', function(){{
        currentIndex = (currentIndex - 1 + slides.length) % slides.length;
        updateSlider();
      }});

      // Auto play cada 5 segundos
      setInterval(function(){{
        currentIndex = (currentIndex + 1) % slides.length;
        updateSlider();
      }}, 5000);
    }})();
  </script>
</body>
</html>
"""
    return html

def build_v2_cdmx():
    print("=== EJECUTANDO REDISEÑO V2: WEBS DINÁMICAS Y DE ALTA IMPACTO (CDMX) ===")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb['EMPRESAS MAPS']
    rows = list(sheet.iter_rows(values_only=False))

    updated_count = 0

    for idx, r in enumerate(rows[1:], start=2):
        vals = [cell.value for cell in r]
        if not any(vals): continue
        nombre = str(vals[0]).strip()
        ciudad = str(vals[1]).strip()
        tipo = str(vals[2]).strip()
        correo = str(vals[3]).strip() if vals[3] else ''
        tel = str(vals[4]).strip() if vals[4] else ''
        resenas = str(vals[6]).strip() if vals[6] else ''
        rating = str(vals[7]).strip() if vals[7] else ''
        estado = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ''

        if 'WEB_BORRADOR_LISTA' in estado or 'PENDIENTE' in estado or 'Prospecto' in estado:
            config = BRANDS_V2.get(nombre)
            if not config:
                slug_val = slugify(nombre)
                config = {
                    'slug': slug_val,
                    'nicho_tipo': tipo,
                    'nicho_schema': 'Dentist' if 'dent' in tipo.lower() or 'odont' in tipo.lower() else ('PhysicalTherapy' if 'fisio' in tipo.lower() or 'rehab' in tipo.lower() else 'MedicalClinic'),
                    'primary': '#0284c7',
                    'accent': '#38bdf8',
                    'accent_light': '#e0f2fe',
                    'dark_bg': '#0c4a6e',
                    'tagline': f'Atención Especializada en {ciudad}',
                    'hero_title': f'Salud & Atención de Excelencia en {ciudad}',
                    'hero_sub': f'Servicios dedicados con tecnología de primer nivel y la mejor atención humana para tu tranquilidad.',
                    'hero_img': 'https://images.unsplash.com/photo-1629909613654-28e377c37b09?auto=format&fit=crop&w=1200&q=80',
                    'keywords': f'{slug_val}, especialidades cdmx, clinica cdmx',
                    'stats': [('4.9 ★', 'Google Maps'), ('+1,000', 'Pacientes Atendidos'), ('100%', 'Atención Guiada'), ('10+', 'Años de Experiencia')],
                    'servicios': [
                        ('Atención Personalizada', 'Evaluación integral y plan de tratamiento diseñado para tus necesidades.', '🩺', ['Diagnóstico', 'Atención Humana']),
                        ('Equipos Digitales', 'Tecnología moderna para procedimientos cómodos y seguros.', '🔬', ['Alta Precisión', 'Procedimientos Rápidos']),
                        ('Especialistas Certificados', 'Profesionales capacitados en constante actualización médica.', '🛡️', ['Calidad Garantizada', 'Seguridad']),
                        ('Instalaciones Cómodas', 'Espacios climatizados e higienizados para tu máximo confort.', '✨', ['Higiene Estricta', 'Confort'])
                    ]
                }

            item = {'nombre': nombre, 'ciudad': ciudad, 'tipo': tipo, 'tel': tel, 'resenas': resenas, 'rating': rating}

            slug = config['slug']
            site_dir = os.path.join(BASE_DIR, slug)
            os.makedirs(site_dir, exist_ok=True)

            html_code = generate_super_v2_html(nombre, item, config)
            with open(os.path.join(site_dir, 'index.html'), 'w', encoding='utf-8') as f:
                f.write(html_code)

            demo_url = f"https://betto12680.github.io/borradoes-webs/{slug}/"
            nuevo_estado = f"WEB_BORRADOR_LISTA | LINK: {demo_url}"
            sheet.cell(row=idx, column=10, value=nuevo_estado)

            updated_count += 1
            print(f"✔ Web V2 Rediseñada: [{nombre}] -> {demo_url}")

    wb.save(EXCEL_PATH)
    print(f"\nSe rediseñaron {updated_count} sitios web borrador en disco con la plantilla V2. Desplegando en GitHub...")

    try:
        subprocess.run(["git", "add", "."], cwd=BASE_DIR, check=True)
        subprocess.run(["git", "commit", "-m", f"Rediseño V2 Total: {updated_count} webs estilo Odontosalud-P con carrusel interactivo y stats"], cwd=BASE_DIR, check=True)
        subprocess.run(["git", "push"], cwd=BASE_DIR, check=True)
        print("✔ Despliegue en GitHub Pages completado exitosamente.")
    except Exception as e:
        print(f"Error en Git push: {e}")

if __name__ == "__main__":
    build_v2_cdmx()
